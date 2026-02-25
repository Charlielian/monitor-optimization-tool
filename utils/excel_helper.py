"""
Excel导出辅助工具
统一处理Excel文件的创建和样式设置
"""
from typing import List, Dict, Any, Optional, Callable, Union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from constants import EXCEL_HEADER_COLOR, EXCEL_HEADER_FONT_COLOR, EXCEL_HEADER_FONT_SIZE, GB_TO_TB
import io
from datetime import datetime


def create_styled_workbook(sheet_name: str = "Sheet1") -> tuple[Workbook, Worksheet]:
    """
    创建带样式的工作簿
    
    Args:
        sheet_name: 工作表名称
    
    Returns:
        (workbook, worksheet) 元组
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws


def apply_header_style(worksheet: Worksheet, row: int = 1) -> None:
    """
    应用表头样式
    
    Args:
        worksheet: 工作表对象
        row: 表头所在行号（默认第1行）
    """
    header_fill = PatternFill(
        start_color=EXCEL_HEADER_COLOR,
        end_color=EXCEL_HEADER_COLOR,
        fill_type="solid"
    )
    header_font = Font(
        bold=True,
        color=EXCEL_HEADER_FONT_COLOR,
        size=EXCEL_HEADER_FONT_SIZE
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in worksheet[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def set_column_widths(worksheet: Worksheet, widths: Dict[str, int]) -> None:
    """
    设置列宽
    
    Args:
        worksheet: 工作表对象
        widths: 列名到宽度的映射字典
    """
    if worksheet.max_row == 0:
        return
    
    # 获取表头
    headers = [cell.value for cell in worksheet[1]]
    
    for idx, header in enumerate(headers, 1):
        if header in widths:
            col_letter = worksheet.cell(row=1, column=idx).column_letter
            worksheet.column_dimensions[col_letter].width = widths[header]


def write_data_to_sheet(
    worksheet: Worksheet,
    headers: List[str],
    data: List[Dict[str, Any]],
    column_widths: Optional[Dict[str, int]] = None
) -> None:
    """
    将数据写入工作表并应用样式
    
    Args:
        worksheet: 工作表对象
        headers: 表头列表
        data: 数据列表
        column_widths: 列宽配置（可选）
    """
    # 写入表头
    worksheet.append(headers)
    
    # 写入数据
    for row in data:
        worksheet.append([row.get(h) for h in headers])
    
    # 应用样式
    apply_header_style(worksheet)
    
    # 设置列宽
    if column_widths:
        set_column_widths(worksheet, column_widths)


def create_template_workbook(
    sheet_name: str,
    headers: List[str],
    sample_data: Optional[List[List[Any]]] = None,
    column_widths: Optional[Dict[str, int]] = None
) -> Workbook:
    """
    创建模板工作簿
    
    Args:
        sheet_name: 工作表名称
        headers: 表头列表
        sample_data: 示例数据（可选）
        column_widths: 列宽配置（可选）
    
    Returns:
        工作簿对象
    """
    wb, ws = create_styled_workbook(sheet_name)
    
    # 写入表头
    ws.append(headers)
    
    # 写入示例数据
    if sample_data:
        for row in sample_data:
            ws.append(row)
    
    # 应用样式
    apply_header_style(ws)
    
    # 设置列宽
    if column_widths:
        set_column_widths(ws, column_widths)

    return wb


def create_alarm_export(
    sheet_name: str,
    headers: List[str],
    data: List[Dict[str, Any]],
    column_widths: Dict[str, int],
    data_mapper: Callable[[Dict[str, Any]], List[Any]],
    filename_prefix: str
) -> io.BytesIO:
    """
    创建告警导出Excel文件（通用函数）

    Args:
        sheet_name: 工作表名称
        headers: 表头列表
        data: 数据列表
        column_widths: 列宽配置
        data_mapper: 数据映射函数，将字典转换为列表
        filename_prefix: 文件名前缀

    Returns:
        BytesIO对象，包含Excel文件内容
    """
    wb, ws = create_styled_workbook(sheet_name)

    # 写入表头
    ws.append(headers)

    # 写入数据
    for row in data:
        ws.append(data_mapper(row))

    # 应用表头样式
    apply_header_style(ws)

    # 设置列宽
    set_column_widths(ws, column_widths)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def format_traffic_value(gb_value: Optional[float], precision: int = 2) -> str:
    """
    智能转换流量单位 (GB/TB)
    
    Args:
        gb_value: 流量值（单位：GB）
        precision: 小数位数（默认2位）
    
    Returns:
        格式化后的字符串，包含单位
    """
    if gb_value is None:
        return ""
    
    try:
        value = float(gb_value)
    except (ValueError, TypeError):
        return str(gb_value)
    
    if value >= GB_TO_TB:
        # 转换为TB
        tb_value = value / GB_TO_TB
        return f"{tb_value:.{precision}f} TB"
    else:
        return f"{value:.{precision}f} GB"


def format_numeric_value(
    value: Any,
    number_format: Optional[str] = None,
    default: Any = ""
) -> Union[int, float, str]:
    """
    格式化数值，用于Excel单元格
    
    Args:
        value: 原始值
        number_format: 数值格式（'int', 'float', 'percent', None）
        default: 值为None时的默认值
    
    Returns:
        格式化后的值
    """
    if value is None:
        return default
    
    if number_format == 'int':
        try:
            return int(value)
        except (ValueError, TypeError):
            return value
    elif number_format in ('float', 'percent'):
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    else:
        return value


def create_export_workbook(
    sheet_name: str,
    headers: List[str],
    data: List[Dict[str, Any]],
    column_widths: Optional[Dict[str, int]] = None,
    data_mapper: Optional[Callable[[Dict[str, Any]], List[Any]]] = None,
    number_formats: Optional[Dict[int, str]] = None,
    freeze_header: bool = True,
    auto_width: bool = False
) -> io.BytesIO:
    """
    创建通用导出工作簿
    
    Args:
        sheet_name: 工作表名称
        headers: 中文表头列表
        data: 数据列表（字典列表）
        column_widths: 列宽配置 {表头名: 宽度}，可选
        data_mapper: 数据映射函数，将字典转换为行数据列表。
                     如果为None，则按headers顺序从字典中取值
        number_formats: 数值格式配置 {列索引(从1开始): 格式字符串}
                       格式字符串如 '0.00', '0', '0.00%' 等
        freeze_header: 是否冻结首行（默认True）
        auto_width: 是否自动计算列宽（默认False，使用column_widths）
    
    Returns:
        BytesIO 对象，包含 Excel 文件内容
    """
    wb, ws = create_styled_workbook(sheet_name)
    
    # 写入表头
    ws.append(headers)
    
    # 写入数据
    for row in data:
        if data_mapper:
            row_data = data_mapper(row)
        else:
            # 默认按headers顺序从字典中取值
            row_data = [row.get(h, "") for h in headers]
        ws.append(row_data)
    
    # 应用表头样式
    apply_header_style(ws)
    
    # 设置数值格式
    if number_formats:
        for row_idx in range(2, ws.max_row + 1):
            for col_idx, fmt in number_formats.items():
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    cell.number_format = fmt
    
    # 设置列宽
    if auto_width:
        _auto_adjust_column_widths(ws)
    elif column_widths:
        set_column_widths(ws, column_widths)
    
    # 冻结首行
    if freeze_header:
        ws.freeze_panes = "A2"
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _auto_adjust_column_widths(worksheet: Worksheet, max_width: int = 50) -> None:
    """
    自动调整列宽
    
    Args:
        worksheet: 工作表对象
        max_width: 最大列宽
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_multi_sheet_workbook(
    sheets_config: List[Dict[str, Any]],
    freeze_header: bool = True
) -> io.BytesIO:
    """
    创建多工作表的导出工作簿
    
    Args:
        sheets_config: 工作表配置列表，每个配置包含：
            - sheet_name: 工作表名称
            - headers: 表头列表
            - data: 数据列表
            - column_widths: 列宽配置（可选）
            - data_mapper: 数据映射函数（可选）
            - number_formats: 数值格式配置（可选）
        freeze_header: 是否冻结首行
    
    Returns:
        BytesIO 对象，包含 Excel 文件内容
    """
    wb = Workbook()
    
    for idx, config in enumerate(sheets_config):
        if idx == 0:
            ws = wb.active
            ws.title = config['sheet_name']
        else:
            ws = wb.create_sheet(title=config['sheet_name'])
        
        headers = config['headers']
        data = config.get('data', [])
        data_mapper = config.get('data_mapper')
        column_widths = config.get('column_widths')
        number_formats = config.get('number_formats')
        
        # 写入表头
        ws.append(headers)
        
        # 写入数据
        for row in data:
            if data_mapper:
                row_data = data_mapper(row)
            else:
                row_data = [row.get(h, "") for h in headers]
            ws.append(row_data)
        
        # 应用表头样式
        apply_header_style(ws)
        
        # 设置数值格式
        if number_formats:
            for row_idx in range(2, ws.max_row + 1):
                for col_idx, fmt in number_formats.items():
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is not None:
                        cell.number_format = fmt
        
        # 设置列宽
        if column_widths:
            set_column_widths(ws, column_widths)
        
        # 冻结首行
        if freeze_header:
            ws.freeze_panes = "A2"
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_export_filename(prefix: str, extension: str = "xlsx") -> str:
    """
    生成带时间戳的导出文件名
    
    Args:
        prefix: 文件名前缀
        extension: 文件扩展名（默认xlsx）
    
    Returns:
        格式化的文件名，如 "前缀_20250110_143000.xlsx"
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"
