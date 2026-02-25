"""
数据导出路由模块
提取自 app.py 的导出相关路由
包括流量导出、监控数据导出、小区数据导出、场景小区导出等功能
"""
import csv
import io
import logging
from datetime import datetime
from typing import List

from flask import Blueprint, request, send_file, flash, redirect, url_for, current_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from utils.excel_helper import (
    create_export_workbook, 
    format_traffic_value,
    create_multi_sheet_workbook,
    generate_export_filename
)
from utils.validators import (
    validate_time_range, 
    validate_granularity,
    validate_and_parse_cgis,
    validate_network_type
)
from utils.time_parser import parse_time_range
from constants import (
    GRANULARITY_15MIN, TOP_CELLS_EXPORT_LIMIT, DEFAULT_THRESHOLD_4G, 
    DEFAULT_THRESHOLD_5G, MAX_CELL_QUERY_LIMIT, FILENAME_DATETIME_FORMAT
)

# 创建导出路由蓝图
export_bp = Blueprint('export', __name__, url_prefix='/export')


@export_bp.route("/traffic.csv")
def export_traffic():
    """导出流量数据为CSV"""
    # 获取服务实例
    service = current_app.config.get('service')
    scenario_service = current_app.config.get('scenario_service')
    cfg = current_app.config.get('app_config')
    
    if not service or not scenario_service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("dashboard"))
    
    range_key = validate_time_range(request.args.get("range", ""))
    networks: List[str] = request.args.getlist("networks") or cfg.ui_config["default_networks"]
    granularity = validate_granularity(request.args.get("granularity", ""))
    
    latest = scenario_service.latest_time()
    latest_ts = max((ts for ts in [latest.get("4g"), latest.get("5g")] if ts), default=None)
    start, end = service.resolve_range(range_key, reference_time=latest_ts)
    data = service.traffic_series(networks, start, end, granularity)
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(data[0].keys()) if data else [])
    writer.writeheader()
    writer.writerows(data)
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    
    return send_file(
        mem,
        as_attachment=True,
        download_name=f"traffic_{range_key}.csv",
        mimetype="text/csv",
    )


@export_bp.route("/top.csv")
def export_top():
    """导出Top小区数据为CSV"""
    service = current_app.config.get('service')
    cfg = current_app.config.get('app_config')
    
    if not service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("dashboard"))
    
    networks: List[str] = request.args.getlist("networks") or cfg.ui_config["default_networks"]
    granularity = validate_granularity(request.args.get("granularity", ""))
    rows = service.top_cells(networks, limit=TOP_CELLS_EXPORT_LIMIT, granularity=granularity)
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()) if rows else [])
    writer.writeheader()
    writer.writerows(rows)
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    
    return send_file(
        mem,
        as_attachment=True,
        download_name="top_cells.csv",
        mimetype="text/csv",
    )


@export_bp.route("/monitor.csv")
def export_monitor():
    """导出保障监控数据为CSV"""
    scenario_service = current_app.config.get('scenario_service')
    
    if not scenario_service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("main.monitor"))
    
    selected = request.args.getlist("scenario_id", type=int)
    thr4 = float(request.args.get("thr4g", DEFAULT_THRESHOLD_4G))
    thr5 = float(request.args.get("thr5g", DEFAULT_THRESHOLD_5G))
    rows = scenario_service.scenario_metrics(selected, thr4, thr5)
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()) if rows else [])
    writer.writeheader()
    writer.writerows(rows)
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    
    return send_file(
        mem,
        as_attachment=True,
        download_name="scenario_monitor.csv",
        mimetype="text/csv",
    )


@export_bp.route("/monitor.xlsx")
def export_monitor_xlsx():
    """导出保障监控数据为Excel"""
    scenario_service = current_app.config.get('scenario_service')
    
    if not scenario_service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("main.monitor"))
    
    selected = request.args.getlist("scenario_id", type=int)
    thr4 = float(request.args.get("thr4g", DEFAULT_THRESHOLD_4G))
    thr5 = float(request.args.get("thr5g", DEFAULT_THRESHOLD_5G))
    rows = scenario_service.scenario_metrics(selected, thr4, thr5)
    
    # 使用通用Excel导出函数
    headers = list(rows[0].keys()) if rows else []
    output = create_export_workbook(
        sheet_name="monitor",
        headers=headers,
        data=rows
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name="scenario_monitor.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def _create_summary_sheet(wb, scenario_service, selected, thr4, thr5, start=None, end=None):
    """创建45G整体汇总sheet"""
    ws1 = wb.create_sheet("45G整体")
    summary_rows = scenario_service.scenario_metrics(selected, thr4, thr5, start, end)
    logging.info(f"Sheet1: 获取到 {len(summary_rows)} 条汇总数据")
    
    if summary_rows:
        chinese_headers = ["场景", "制式", "时间", "流量值", "流量单位", "上行PRB利用率(%)",
                         "下行PRB利用率(%)", "无线接通率(%)", "最大用户数", "超阈值小区数", "总小区数"]
        ws1.append(chinese_headers)
        for r in summary_rows:
            ws1.append([
                r.get("scenario", ""),
                r.get("network", ""),  # 修正：使用 "network" 而不是 "network_type"
                str(r.get("ts", "")),  # 修正：使用 "ts" 而不是 "time"
                r.get("流量值", 0),
                r.get("流量单位", ""),
                r.get("上行PRB利用率(%)", 0),  # 修正：添加括号
                r.get("下行PRB利用率(%)", 0),  # 修正：添加括号
                r.get("无线接通率(%)", 0),     # 修正：添加括号
                r.get("最大用户数", 0),
                r.get("超阈值小区数", 0),
                r.get("总小区数", 0),
            ])
    return ws1


@export_bp.route("/monitor_xlsx_full")
def export_monitor_xlsx_full():
    """导出保障监控数据：3个sheet（45G整体、4G小区、5G小区）"""
    scenario_service = current_app.config.get('scenario_service')
    
    if not scenario_service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("main.monitor"))
    
    try:
        selected = request.args.getlist("scenario_id", type=int)
        thr4 = float(request.args.get("thr4g", DEFAULT_THRESHOLD_4G))
        thr5 = float(request.args.get("thr5g", DEFAULT_THRESHOLD_5G))

        logging.info(f"开始导出保障监控数据，场景ID: {selected}, 阈值4G: {thr4}, 阈值5G: {thr5}")

        # 时间解析（使用统一工具函数）
        latest = scenario_service.latest_time()
        latest_ts_candidates = [latest.get("4g"), latest.get("5g")]
        latest_ts = max((ts for ts in latest_ts_candidates if ts), default=None)

        start, end = parse_time_range(
            request.args.get("start_time", ""),
            request.args.get("end_time", ""),
            latest_ts=latest_ts,
            default_hours=6,
            max_days=30
        )

        logging.info(f"时间范围: {start} 到 {end}")

        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        # Sheet1: 45G整体（场景汇总指标）- 使用时间范围
        ws1 = _create_summary_sheet(wb, scenario_service, selected, thr4, thr5, start, end)

        # Sheet2: 4G小区指标
        ws2 = wb.create_sheet("4G小区指标")
        # 批量获取所有场景的小区数据（解决N+1查询问题）
        cells_by_scenario = scenario_service.list_cells_batch(selected)
        cells_4g_all = []

        for sid in selected:
            s_name = scenario_service.get_scenario_name(sid)
            if not s_name:
                logging.warning(f"场景ID {sid} 未找到场景名称")
                continue
            cells = [c for c in cells_by_scenario.get(sid, []) if c["network_type"] == "4G"]
            logging.info(f"场景 {s_name} (ID:{sid}) 有 {len(cells)} 个4G小区")
            for c in cells:
                cells_4g_all.append({"scenario": s_name, **c})
        
        logging.info(f"Sheet2: 共收集到 {len(cells_4g_all)} 个4G小区")
        
        if cells_4g_all:
            # 创建小区映射
            cell_map_4g = {}
            for c in cells_4g_all:
                cgi = c.get("cgi", "")
                if cgi:
                    cell_map_4g[cgi] = {
                        "scenario": c.get("scenario", ""),
                        "cell_id": c.get("cell_id", ""),
                        "cell_name": c.get("cell_name", ""),
                        "cgi": cgi,
                    }
            
            cgis_4g = list(cell_map_4g.keys())
            logging.info(f"Sheet2: 有效CGI数量: {len(cgis_4g)}")
            
            # 查询时间范围内的所有数据
            metrics_map_4g = {}
            if cgis_4g:
                sql_4g = f"""
                    SELECT 
                        cell_id, cgi, start_time, cellname,
                        (COALESCE("PDCP_UpOctUl",0) + COALESCE("PDCP_UpOctDl",0)) / 1000.0 / 1000.0 AS traffic_gb,
                        ul_prb_utilization AS ul_prb_util,
                        dl_prb_utilization AS dl_prb_util,
                        GREATEST(COALESCE(ul_prb_utilization,0), COALESCE(dl_prb_utilization,0)) AS max_prb_util,
                        wireless_connect_rate AS connect_rate,
                        "RRC_ConnMax" AS rrc_users,
                        interference
                    FROM cell_4g_metrics
                    WHERE start_time BETWEEN %s AND %s AND cgi IN ({','.join(['%s']*len(cgis_4g))})
                    ORDER BY start_time DESC, max_prb_util DESC
                """
                params = [start, end] + cgis_4g
                rows_4g = scenario_service.pg.fetch_all(sql_4g, tuple(params)) or []
                logging.info(f"Sheet2: 查询到 {len(rows_4g)} 条4G指标数据")
            
            # 建立指标映射（CGI -> 指标列表）
            for row in rows_4g:
                cgi = row.get("cgi", "")
                if cgi:
                    if cgi not in metrics_map_4g:
                        metrics_map_4g[cgi] = []
                    metrics_map_4g[cgi].append(row)
        
        # 写入Sheet2
        chinese_headers_4g = ["场景", "小区ID", "CGI", "小区名", "时间", "流量(GB)", 
                             "上行PRB利用率(%)", "下行PRB利用率(%)", "最大PRB利用率(%)", 
                             "无线接通率(%)", "最大用户数", "干扰", "数据状态"]
        ws2.append(chinese_headers_4g)
        
        # 合并小区信息和指标数据
        for cgi, cell_info in cell_map_4g.items():
            if cgi in metrics_map_4g:
                # 有数据的小区，输出所有时间点的数据
                for row in metrics_map_4g[cgi]:
                    traffic_gb = float(row.get("traffic_gb") or 0)
                    ws2.append([
                        cell_info["scenario"],
                        row.get("cell_id", cell_info["cell_id"]),
                        cgi,
                        row.get("cellname", cell_info["cell_name"]),
                        str(row.get("start_time", "")),
                        f"{traffic_gb:.2f}",
                        f"{float(row.get('ul_prb_util') or 0):.2f}",
                        f"{float(row.get('dl_prb_util') or 0):.2f}",
                        f"{float(row.get('max_prb_util') or 0):.2f}",
                        f"{float(row.get('connect_rate') or 0):.2f}",
                        int(row.get("rrc_users") or 0),
                        f"{float(row.get('interference') or 0):.2f}",
                        "有数据",
                    ])
            else:
                # 没有数据的小区，输出一行空值
                ws2.append([
                    cell_info["scenario"],
                    cell_info["cell_id"],
                    cgi,
                    cell_info["cell_name"],
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                    "无数据",
                ])
        
        # Sheet3: 5G小区指标
        ws3 = wb.create_sheet("5G小区指标")
        cells_5g_all = []
        # 复用之前批量获取的数据
        for sid in selected:
            s_name = scenario_service.get_scenario_name(sid)
            if not s_name:
                logging.warning(f"场景ID {sid} 未找到场景名称")
                continue
            cells = [c for c in cells_by_scenario.get(sid, []) if c["network_type"] == "5G"]
            logging.info(f"场景 {s_name} (ID:{sid}) 有 {len(cells)} 个5G小区")
            for c in cells:
                cells_5g_all.append({"scenario": s_name, **c})
        
        if cells_5g_all:
            # 创建小区映射
            cell_map_5g = {}
            for c in cells_5g_all:
                cgi = c.get("cgi", "")
                if cgi:
                    cell_map_5g[cgi] = {
                        "scenario": c.get("scenario", ""),
                        "cell_id": c.get("cell_id", ""),
                        "cell_name": c.get("cell_name", ""),
                        "cgi": cgi,
                    }
            
            cgis_5g = list(cell_map_5g.keys())
            logging.info(f"Sheet3: 有效CGI数量: {len(cgis_5g)}")
            
            # 查询时间范围内的所有数据
            metrics_map_5g = {}
            if cgis_5g:
                sql_5g = f"""
                    SELECT 
                        "Ncgi" AS cell_id, "Ncgi" AS cgi, start_time, userlabel AS cellname,
                        (COALESCE("RLC_UpOctUl",0) + COALESCE("RLC_UpOctDl",0)) / 1000.0 / 1000.0 AS traffic_gb,
                        "RRU_PuschPrbAssn" * 100.0 / NULLIF("RRU_PuschPrbTot", 0) AS ul_prb_util,
                        "RRU_PdschPrbAssn" * 100.0 / NULLIF("RRU_PdschPrbTot", 0) AS dl_prb_util,
                        CASE
                            WHEN ("RRU_PuschPrbAssn" * 100.0 / NULLIF("RRU_PuschPrbTot", 0)) >
                                 ("RRU_PdschPrbAssn" * 100.0 / NULLIF("RRU_PdschPrbTot", 0))
                            THEN ("RRU_PuschPrbAssn" * 100.0 / NULLIF("RRU_PuschPrbTot", 0))
                            ELSE ("RRU_PdschPrbAssn" * 100.0 / NULLIF("RRU_PdschPrbTot", 0))
                        END AS max_prb_util,
                        ("RRC_SuccConnEstab" * 100.0 / NULLIF("RRC_AttConnEstab", 0)) *
                        ("NGSIG_ConnEstabSucc" * 100.0 / NULLIF("NGSIG_ConnEstabAtt", 0)) *
                        ("Flow_NbrSuccEstab" * 100.0 / NULLIF("Flow_NbrAttEstab", 0)) / 100.0 / 100.0 AS connect_rate,
                        "RRC_ConnMax" AS rrc_users,
                        interference
                    FROM cell_5g_metrics
                    WHERE start_time BETWEEN %s AND %s AND "Ncgi" IN ({','.join(['%s']*len(cgis_5g))})
                    ORDER BY start_time DESC, max_prb_util DESC
                """
                params_5g = [start, end] + cgis_5g
                rows_5g = scenario_service.pg.fetch_all(sql_5g, tuple(params_5g)) or []
                logging.info(f"Sheet3: 查询到 {len(rows_5g)} 条5G指标数据")
                
                # 建立指标映射（CGI -> 指标列表）
                for row in rows_5g:
                    cgi = row.get("cgi", "")
                    if cgi:
                        if cgi not in metrics_map_5g:
                            metrics_map_5g[cgi] = []
                        metrics_map_5g[cgi].append(row)
            
            # 写入Sheet3
            chinese_headers_5g = ["场景", "小区ID", "CGI", "小区名", "时间", "流量(GB)", 
                                 "上行PRB利用率(%)", "下行PRB利用率(%)", "最大PRB利用率(%)", 
                                 "无线接通率(%)", "最大用户数", "干扰", "数据状态"]
            ws3.append(chinese_headers_5g)
            
            # 合并小区信息和指标数据
            for cgi, cell_info in cell_map_5g.items():
                if cgi in metrics_map_5g:
                    # 有数据的小区，输出所有时间点的数据
                    for row in metrics_map_5g[cgi]:
                        traffic_gb = float(row.get("traffic_gb") or 0)
                        ws3.append([
                            cell_info["scenario"],
                            row.get("cell_id", cell_info["cell_id"]),
                            cgi,
                            row.get("cellname", cell_info["cell_name"]),
                            str(row.get("start_time", "")),
                            f"{traffic_gb:.2f}",
                            f"{float(row.get('ul_prb_util') or 0):.2f}",
                            f"{float(row.get('dl_prb_util') or 0):.2f}",
                            f"{float(row.get('max_prb_util') or 0):.2f}",
                            f"{float(row.get('connect_rate') or 0):.2f}",
                            int(row.get("rrc_users") or 0),
                            f"{float(row.get('interference') or 0):.2f}",
                            "有数据",
                        ])
                else:
                    # 没有数据的小区，输出一行空值
                    ws3.append([
                        cell_info["scenario"],
                        cell_info["cell_id"],
                        cgi,
                        cell_info["cell_name"],
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "无数据",
                    ])
    
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for ws in [ws1, ws2, ws3]:
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
        
        logging.info("Excel文件生成成功，准备返回")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"保障监控数据_{start.strftime(FILENAME_DATETIME_FORMAT)}_{end.strftime(FILENAME_DATETIME_FORMAT)}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logging.error(f"导出保障监控数据失败: {str(e)}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("main.monitor"))


@export_bp.route("/latest_metrics.xlsx")
def export_latest_metrics():
    """导出最新全量小区指标为Excel，使用中文列名"""
    service = current_app.config.get('service')
    
    if not service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("dashboard"))
    
    rows = service.latest_full_metrics()
    
    if not rows:
        flash("暂无数据可导出", "warning")
        return redirect(url_for("dashboard"))
    
    # 定义中文列名映射
    column_mapping = {
        "network": "网络类型",
        "cellname": "小区名",
        "cell_id": "小区ID",
        "cgi": "CGI",
        "start_time": "时间",
        "total_traffic_gb": "总流量(GB)",
        "ul_prb_utilization": "上行PRB利用率(%)",
        "dl_prb_utilization": "下行PRB利用率(%)",
        "wireless_connect_rate": "无线接通率(%)",
        "rrc_users": "最大RRC连接数",
        "interference": "干扰(dBm)"
    }
    
    # 数据映射函数
    def data_mapper(row):
        row_data = []
        for eng_key, cn_key in column_mapping.items():
            value = row.get(eng_key)
            # 格式化数值
            if value is not None:
                if eng_key in ["total_traffic_gb", "ul_prb_utilization", "dl_prb_utilization", "wireless_connect_rate", "interference"]:
                    try:
                        row_data.append(float(value))
                    except (ValueError, TypeError):
                        row_data.append(value)
                elif eng_key == "rrc_users":
                    try:
                        row_data.append(int(value))
                    except (ValueError, TypeError):
                        row_data.append(value)
                elif eng_key == "start_time":
                    row_data.append(str(value))
                else:
                    row_data.append(value)
            else:
                row_data.append("")
        return row_data
    
    # 列宽配置
    column_widths = {
        "网络类型": 12,
        "小区名": 30,
        "小区ID": 20,
        "CGI": 25,
        "时间": 20,
        "总流量(GB)": 15,
        "上行PRB利用率(%)": 18,
        "下行PRB利用率(%)": 18,
        "无线接通率(%)": 15,
        "最大RRC连接数": 15,
        "干扰(dBm)": 12
    }
    
    # 数值格式配置
    number_formats = {
        6: '0.00',   # 总流量(GB)
        7: '0.00',   # 上行PRB利用率(%)
        8: '0.00',   # 下行PRB利用率(%)
        9: '0.00',   # 无线接通率(%)
        11: '0.00',  # 干扰(dBm)
    }
    
    # 使用通用Excel导出函数
    output = create_export_workbook(
        sheet_name="最新全量小区指标",
        headers=list(column_mapping.values()),
        data=rows,
        column_widths=column_widths,
        data_mapper=data_mapper,
        number_formats=number_formats
    )
    
    # 生成文件名（包含时间戳）
    filename = generate_export_filename("最新全量小区指标")
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@export_bp.route("/cell_data.xlsx")
def export_cell_data():
    """导出指标查询结果为xlsx，列名使用中文"""
    service = current_app.config.get('service')
    scenario_service = current_app.config.get('scenario_service')
    
    if not service or not scenario_service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("cell"))
    
    cell_cgi = request.args.get("cell_cgi", "").strip()
    cell_network = request.args.get("cell_network", "").strip()
    granularity = validate_granularity(request.args.get("granularity", ""))
    query_type = request.args.get("query_type", "cell").strip()
    scenario_ids = request.args.getlist("scenario_ids")
    include_metrics = request.args.get("include_metrics", "0").strip() == "1"

    # 从场景中获取小区列表（如果选择了场景）
    selected_scenario_ids = []
    for sid in scenario_ids:
        if sid and sid.isdigit():
            selected_scenario_ids.append(int(sid))

    if selected_scenario_ids:
        all_scenario_cgis = []
        for scenario_id in selected_scenario_ids:
            scenario_cells = scenario_service.list_cells(scenario_id)
            scenario_cgis = [cell.get('cgi') or cell.get('cell_id') for cell in scenario_cells if cell.get('cgi') or cell.get('cell_id')]
            all_scenario_cgis.extend(scenario_cgis)
        if all_scenario_cgis:
            cell_cgi = ','.join(all_scenario_cgis)

    # 时间解析（使用统一工具函数）
    latest = scenario_service.latest_time()
    latest_ts_candidates = [latest.get("4g"), latest.get("5g")]
    latest_ts = max((ts for ts in latest_ts_candidates if ts), default=None)
    
    # 根据粒度设置默认时间范围
    if granularity == "1d":
        default_hours = 168  # 天级默认7天
        max_days = 90  # 天级最多90天
    else:
        default_hours = 6  # 小时级默认6小时
        max_days = 30  # 小时级最多30天
    
    start, end = parse_time_range(
        request.args.get("start_time", ""),
        request.args.get("end_time", ""),
        latest_ts=latest_ts,
        default_hours=default_hours,
        max_days=max_days
    )

    # CGI验证和查询数据
    cell_data = []
    if cell_cgi:
        cgi_list, _ = validate_and_parse_cgis(cell_cgi, MAX_CELL_QUERY_LIMIT)
        
        # 如果是汇总查询
        if query_type == "summary":
            # 如果未指定制式，自动识别（混查模式）
            if not cell_network or cell_network == "混查":
                # 自动识别CGI并分别汇总
                cgi_4g = []
                cgi_5g = []
                for cgi in cgi_list:
                    if not cgi:
                        continue
                    parts = cgi.split('-')
                    if len(parts) >= 4:
                        enb_gnb = parts[2]
                        if len(enb_gnb) == 6:
                            cgi_4g.append(cgi)
                        elif len(enb_gnb) == 8:
                            cgi_5g.append(cgi)
                
                # 分别汇总4G和5G数据
                summary_4g = []
                summary_5g = []
                
                if cgi_4g:
                    summary_4g = service.cell_timeseries_summary(cgi_4g, "4G", start, end, granularity)
                if cgi_5g:
                    summary_5g = service.cell_timeseries_summary(cgi_5g, "5G", start, end, granularity)
                
                # 合并结果
                cell_data = summary_4g + summary_5g
                # 按时间排序
                cell_data.sort(key=lambda x: x.get('start_time', ''))
            else:
                # 指定制式汇总查询
                cell_network = validate_network_type(cell_network)
                cell_data = service.cell_timeseries_summary(
                    cgi_list, cell_network, start, end, granularity
                )
        else:
            # 小区查询
            # 如果是混查模式
            if not cell_network or cell_network == "混查":
                cell_data, _ = service.cell_timeseries_mixed(cgi_list, start, end, granularity)
            else:
                cell_network = validate_network_type(cell_network)
                cell_data = service.cell_timeseries_bulk(cgi_list, cell_network, start, end, granularity)
                for row in cell_data:
                    if 'cgi' not in row or not row.get('cgi'):
                        row['cgi'] = row.get('cell_id')
                    if 'cell_id' not in row or not row.get('cell_id'):
                        row['cell_id'] = row.get('cgi')
                    row['network_type'] = cell_network

        # 如果需要导出分子分母，查询并添加真实的分子分母数据
        if include_metrics and cell_data:
            # 创建一个映射表，用于快速查找和更新数据
            # 键为 (cgi, start_time)，值为 cell_data 中的行索引
            data_map = {}
            for idx, row in enumerate(cell_data):
                key = (row.get('cgi') or row.get('cell_id'), str(row.get('start_time')))
                data_map[key] = idx

            # 分别查询4G和5G的分子分母数据
            # 收集所有需要查询的CGI
            cgis_4g = []
            cgis_5g = []
            for row in cell_data:
                cgi = row.get('cgi') or row.get('cell_id')
                if not cgi:
                    continue
                if row.get('network_type') == '4G':
                    cgis_4g.append(cgi)
                elif row.get('network_type') == '5G':
                    cgis_5g.append(cgi)

            # 查询4G分子分母数据
            if cgis_4g:
                table_4g = service.get_table_name("4G", granularity)
                placeholders = ','.join(['%s'] * len(cgis_4g))
                sql_4g = f"""
                    SELECT
                        start_time,
                        cgi,
                        COALESCE("RRC_SuccConnEstab", 0) AS rrc_success_count,
                        COALESCE("RRC_AttConnEstab", 0) AS rrc_attempt_count,
                        COALESCE("ERAB_NbrSuccEstab", 0) AS erab_success_count,
                        COALESCE("ERAB_NbrAttEstab", 0) AS erab_attempt_count,
                        (COALESCE("HO_SuccOutInterEnbS1",0) + COALESCE("HO_SuccOutInterEnbX2",0) + COALESCE("HO_SuccOutIntraEnb",0)) AS ho_success_count,
                        (COALESCE("HO_AttOutInterEnbS1",0) + COALESCE("HO_AttOutInterEnbX2",0) + COALESCE("HO_AttOutIntraEnb",0)) AS ho_attempt_count,
                        (COALESCE("ERAB_HoFail",0) - COALESCE("ERAB_NbrReqRelEnb_Normal",0) + COALESCE("ERAB_NbrReqRelEnb",0)) AS drop_numerator,
                        (COALESCE("CONTEXT_NbrLeft",0) + COALESCE("ERAB_NbrSuccEstab",0) + COALESCE("ERAB_NbrHoInc",0)) AS drop_denominator
                    FROM {table_4g}
                    WHERE start_time BETWEEN %s AND %s
                      AND cgi IN ({placeholders})
                """
                params_4g = [start, end] + cgis_4g
                metrics_4g = service.pg.fetch_all(sql_4g, tuple(params_4g)) or []

                # 将查询结果合并到 cell_data
                for metric in metrics_4g:
                    key = (metric.get('cgi'), str(metric.get('start_time')))
                    if key in data_map:
                        idx = data_map[key]
                        cell_data[idx].update({
                            'rrc_success_count': metric.get('rrc_success_count', 0),
                            'rrc_attempt_count': metric.get('rrc_attempt_count', 0),
                            'erab_success_count': metric.get('erab_success_count', 0),
                            'erab_attempt_count': metric.get('erab_attempt_count', 0),
                            'ho_success_count': metric.get('ho_success_count', 0),
                            'ho_attempt_count': metric.get('ho_attempt_count', 0),
                            'drop_numerator': metric.get('drop_numerator', 0),
                            'drop_denominator': metric.get('drop_denominator', 0)
                        })

            # 查询5G分子分母数据
            if cgis_5g:
                table_5g = service.get_table_name("5G", granularity)
                placeholders = ','.join(['%s'] * len(cgis_5g))
                sql_5g = f"""
                    SELECT
                        start_time,
                        "Ncgi" AS cgi,
                        COALESCE("RRC_SuccConnEstab", 0) AS rrc_success_count,
                        COALESCE("RRC_AttConnEstab", 0) AS rrc_attempt_count,
                        COALESCE("NGSIG_ConnEstabSucc", 0) AS ng_success_count,
                        COALESCE("NGSIG_ConnEstabAtt", 0) AS ng_attempt_count,
                        COALESCE("Flow_NbrSuccEstab", 0) AS flow_success_count,
                        COALESCE("Flow_NbrAttEstab", 0) AS flow_attempt_count,
                        (COALESCE("HO_SuccOutIntraFreq",0) + COALESCE("HO_SuccOutInterFreq",0)) AS ho_success_count,
                        (COALESCE("HO_AttOutExecIntraFreq",0) + COALESCE("HO_AttOutExecInterFreq",0)) AS ho_attempt_count,
                        (COALESCE("CONTEXT_AttRelgNB",0) - COALESCE("CONTEXT_AttRelgNB_Normal",0)) AS drop_numerator,
                        (COALESCE("CONTEXT_SuccInitalSetup",0) + COALESCE("CONTEXT_NbrLeft",0) + COALESCE("HO_SuccExecInc",0) + COALESCE("RRC_SuccConnReestab_NonSrccell",0)) AS drop_denominator
                    FROM {table_5g}
                    WHERE start_time BETWEEN %s AND %s
                      AND "Ncgi" IN ({placeholders})
                """
                params_5g = [start, end] + cgis_5g
                metrics_5g = service.pg.fetch_all(sql_5g, tuple(params_5g)) or []

                # 将查询结果合并到 cell_data
                for metric in metrics_5g:
                    key = (metric.get('cgi'), str(metric.get('start_time')))
                    if key in data_map:
                        idx = data_map[key]
                        cell_data[idx].update({
                            'rrc_success_count': metric.get('rrc_success_count', 0),
                            'rrc_attempt_count': metric.get('rrc_attempt_count', 0),
                            'ng_success_count': metric.get('ng_success_count', 0),
                            'ng_attempt_count': metric.get('ng_attempt_count', 0),
                            'flow_success_count': metric.get('flow_success_count', 0),
                            'flow_attempt_count': metric.get('flow_attempt_count', 0),
                            'ho_success_count': metric.get('ho_success_count', 0),
                            'ho_attempt_count': metric.get('ho_attempt_count', 0),
                            'drop_numerator': metric.get('drop_numerator', 0),
                            'drop_denominator': metric.get('drop_denominator', 0)
                        })

    # 中文列名（包含新指标）
    chinese_headers = [
        "时间", "小区CGI", "小区名", "干扰", 
        "上行PRB利用率(%)", "下行PRB利用率(%)", "最大PRB利用率(%)", 
        "无线接通率(%)", "切换成功率(%)", "掉线率(%)",
        "流量(GB)", "RRC最大用户数", "话务量(Erl)"
    ]
    
    # 如果需要导出分子分母，添加分子分母列
    if include_metrics:
        chinese_headers.extend([
            "RRC成功数", "RRC尝试数",
            "ERAB成功数", "ERAB尝试数",  # 4G专用
            "NG成功数", "NG尝试数",      # 5G专用
            "Flow成功数", "Flow尝试数",  # 5G专用
            "切换成功数", "切换尝试数",
            "掉线分子", "掉线分母"
        ])
    
    # 数据映射函数
    def data_mapper(row):
        traffic_gb = float(row.get('total_traffic', 0))
        ul_prb = float(row.get('ul_prb_utilization', 0) or 0)
        dl_prb = float(row.get('dl_prb_utilization', 0) or 0)
        max_prb = ul_prb if ul_prb > dl_prb else dl_prb
        rrc_users = int(row.get('rrc_users', 0) or 0)
        voice_erl = float(row.get('voice_erl', 0) or 0)
        ho_success_rate = float(row.get('ho_success_rate', 0) or 0)
        
        # 处理掉线率（根据网络类型）
        if row.get('erab_drop_rate') is not None:
            drop_rate = float(row.get('erab_drop_rate') or 0)
        elif row.get('wireless_drop_rate_5g') is not None:
            drop_rate = float(row.get('wireless_drop_rate_5g') or 0)
        else:
            drop_rate = 0
        
        result = [
            row.get('start_time', ''),
            row.get('cgi') or row.get('cell_id', ''),
            row.get('cellname', ''),
            float(row.get('interference', 0) or 0),
            ul_prb,
            dl_prb,
            max_prb,
            float(row.get('wireless_connect_rate', 0) or 0),
            ho_success_rate,
            drop_rate,
            traffic_gb,
            rrc_users,
            voice_erl,
        ]
        
        # 如果需要导出分子分母，添加分子分母数据
        if include_metrics:
            # 根据网络类型添加分子分母数据
            if row.get('network_type') == '4G':
                result.extend([
                    row.get('rrc_success_count', 0),   # RRC成功数
                    row.get('rrc_attempt_count', 0),   # RRC尝试数
                    row.get('erab_success_count', 0),  # ERAB成功数
                    row.get('erab_attempt_count', 0),  # ERAB尝试数
                    0,  # NG成功数 (5G专用)
                    0,  # NG尝试数 (5G专用)
                    0,  # Flow成功数 (5G专用)
                    0,  # Flow尝试数 (5G专用)
                    row.get('ho_success_count', 0),    # 切换成功数
                    row.get('ho_attempt_count', 0),    # 切换尝试数
                    row.get('drop_numerator', 0),      # 掉线分子
                    row.get('drop_denominator', 0)     # 掉线分母
                ])
            elif row.get('network_type') == '5G':
                result.extend([
                    row.get('rrc_success_count', 0),   # RRC成功数
                    row.get('rrc_attempt_count', 0),   # RRC尝试数
                    0,  # ERAB成功数 (4G专用)
                    0,  # ERAB尝试数 (4G专用)
                    row.get('ng_success_count', 0),    # NG成功数
                    row.get('ng_attempt_count', 0),    # NG尝试数
                    row.get('flow_success_count', 0),  # Flow成功数
                    row.get('flow_attempt_count', 0),  # Flow尝试数
                    row.get('ho_success_count', 0),    # 切换成功数
                    row.get('ho_attempt_count', 0),    # 切换尝试数
                    row.get('drop_numerator', 0),      # 掉线分子
                    row.get('drop_denominator', 0)     # 掉线分母
                ])
            else:
                result.extend([
                    0,  # RRC成功数
                    0,  # RRC尝试数
                    0,  # ERAB成功数
                    0,  # ERAB尝试数
                    0,  # NG成功数
                    0,  # NG尝试数
                    0,  # Flow成功数
                    0,  # Flow尝试数
                    0,  # 切换成功数
                    0,  # 切换尝试数
                    0,  # 掉线分子
                    0   # 掉线分母
                ])
        
        return result
    
    # 列宽配置
    column_widths = {
        "时间": 20,
        "小区CGI": 25,
        "小区名": 30,
        "干扰": 12,
        "上行PRB利用率(%)": 18,
        "下行PRB利用率(%)": 18,
        "最大PRB利用率(%)": 18,
        "无线接通率(%)": 15,
        "切换成功率(%)": 15,
        "掉线率(%)": 12,
        "流量(GB)": 15,
        "RRC最大用户数": 15,
        "话务量(Erl)": 15,
    }
    
    # 如果需要导出分子分母，添加分子分母列的列宽
    if include_metrics:
        column_widths.update({
            "RRC成功数": 12,
            "RRC尝试数": 12,
            "ERAB成功数": 12,
            "ERAB尝试数": 12,
            "NG成功数": 12,
            "NG尝试数": 12,
            "Flow成功数": 12,
            "Flow尝试数": 12,
            "切换成功数": 12,
            "切换尝试数": 12,
            "掉线分子": 12,
            "掉线分母": 12,
        })
    
    # 数值格式配置
    number_formats = {
        4: '0.00',   # 干扰
        5: '0.00',   # 上行PRB利用率
        6: '0.00',   # 下行PRB利用率
        7: '0.00',   # 最大PRB利用率
        8: '0.00',   # 无线接通率
        9: '0.00',   # 切换成功率
        10: '0.00',  # 掉线率
        11: '0.00',  # 流量
        13: '0.00',  # 话务量
    }
    
    # 如果需要导出分子分母，添加分子分母列的数值格式
    if include_metrics:
        number_formats.update({
            14: '0',   # RRC成功数
            15: '0',   # RRC尝试数
            16: '0',   # ERAB成功数
            17: '0',   # ERAB尝试数
            18: '0',   # NG成功数
            19: '0',   # NG尝试数
            20: '0',   # Flow成功数
            21: '0',   # Flow尝试数
            22: '0',   # 切换成功数
            23: '0',   # 切换尝试数
            24: '0',   # 掉线分子
            25: '0'    # 掉线分母
        })
    
    # 使用通用Excel导出函数
    output = create_export_workbook(
        sheet_name=f"{query_type == 'summary' and '汇总' or cell_network}小区指标",
        headers=chinese_headers,
        data=cell_data,
        column_widths=column_widths,
        data_mapper=data_mapper,
        number_formats=number_formats
    )
    
    # 生成文件名
    filename = f"指标查询_{query_type == 'summary' and '汇总' or cell_network}_{start.strftime(FILENAME_DATETIME_FORMAT)}_{end.strftime(FILENAME_DATETIME_FORMAT)}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@export_bp.route("/top_utilization.xlsx")
def export_top_utilization():
    """导出45G利用率Top清单为xlsx，列名使用中文"""
    service = current_app.config.get('service')
    
    if not service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("dashboard"))
    
    network = validate_network_type(request.args.get("network", ""))
    limit = int(request.args.get("limit", TOP_CELLS_EXPORT_LIMIT))
    granularity = validate_granularity(request.args.get("granularity", ""))
    rows = service.top_utilization(network, limit=limit, granularity=granularity)
    
    # 中文列名映射
    column_mapping = {
        "cell_id": "小区ID",
        "cellname": "小区名",
        "cgi": "CGI",
        "network_type": "制式",
        "total_traffic": "流量(GB)",
        "ul_prb_util": "上行PRB利用率(%)",
        "dl_prb_util": "下行PRB利用率(%)",
        "max_prb_util": "最大PRB利用率(%)",
        "connect_rate": "无线接通率(%)",
        "max_rrc_users": "最大用户数",
        "interference": "干扰"
    }
    
    # 数据映射函数
    def data_mapper(row):
        row_data = []
        original_headers = list(row.keys())
        for h in original_headers:
            value = row.get(h)
            if h == "total_traffic" and value:
                traffic_gb = float(value)
                if traffic_gb >= 1024:
                    value = f"{traffic_gb / 1024:.2f} TB"
                else:
                    value = f"{traffic_gb:.2f} GB"
            elif h in ["ul_prb_util", "dl_prb_util", "max_prb_util", "connect_rate", "interference"]:
                value = f"{float(value or 0):.2f}" if value is not None else "0.00"
            row_data.append(value)
        return row_data
    
    # 列宽配置
    column_widths = {
        "小区ID": 20,
        "小区名": 25,
        "CGI": 25,
        "制式": 10,
        "流量(GB)": 15,
        "上行PRB利用率(%)": 18,
        "下行PRB利用率(%)": 18,
        "最大PRB利用率(%)": 18,
        "无线接通率(%)": 15,
        "最大用户数": 12,
        "干扰": 12
    }
    
    if rows:
        # 获取原始列名并转换为中文列名
        original_headers = list(rows[0].keys())
        chinese_headers = [column_mapping.get(h, h) for h in original_headers]
        
        # 使用通用Excel导出函数
        output = create_export_workbook(
            sheet_name=f"{network}利用率Top清单",
            headers=chinese_headers,
            data=rows,
            column_widths=column_widths,
            data_mapper=data_mapper
        )
    else:
        # 空数据情况
        output = create_export_workbook(
            sheet_name=f"{network}利用率Top清单",
            headers=list(column_mapping.values()),
            data=[]
        )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{network}利用率Top清单.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/dashboard_data.xlsx")
def export_dashboard_data():
    """导出仪表盘数据为Excel，根据时间（粒度及周期）控制框相关的数据"""
    service = current_app.config.get('service')
    scenario_service = current_app.config.get('scenario_service')
    cfg = current_app.config.get('app_config')
    
    if not service or not scenario_service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("dashboard"))
    
    # 获取并验证参数
    range_key = validate_time_range(request.args.get("range", ""))
    networks: List[str] = request.args.getlist("networks") or cfg.ui_config["default_networks"]
    granularity = validate_granularity(request.args.get("granularity", ""))
    
    # 获取最新时间并解析时间范围
    latest = scenario_service.latest_time()
    latest_ts = max((ts for ts in [latest.get("4g"), latest.get("5g")] if ts), default=None)
    start, end = service.resolve_range(range_key, reference_time=latest_ts)
    
    # 查询数据
    traffic_data = service.traffic_series(networks, start, end, granularity)
    rrc_data = service.rrc_series(networks, start, end, granularity)
    connect_data = service.connectivity_series(networks, start, end, granularity)
    voice_data = service.voice_series(networks, start, end, granularity)
    
    # 创建Excel工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # Sheet1: 全网数据（无线接通，流量，用户数）
    ws1 = wb.create_sheet("全网数据")
    
    # 全网数据列名
    network_headers = ["时间", "网络类型", "无线接通率(%)", "总流量(GB)", "上行PRB利用率(%)", "下行PRB利用率(%)", "话务量(Erl)", "RRC最大用户数"]
    ws1.append(network_headers)
    
    # 合并数据
    # 按时间和网络类型分组
    merged_data = {}
    
    # 处理流量数据
    for row in traffic_data:
        key = (str(row.get("start_time", "")), row.get("network_type", ""))
        if key not in merged_data:
            merged_data[key] = {"traffic": 0, "connect_rate": 0, "rrc_users": 0, "ul_prb": 0, "dl_prb": 0, "voice_erl": 0}
        merged_data[key]["traffic"] = float(row.get("total_traffic") or 0)
    
    # 处理RRC用户数数据
    for row in rrc_data:
        key = (str(row.get("start_time", "")), row.get("network_type", ""))
        if key not in merged_data:
            merged_data[key] = {"traffic": 0, "connect_rate": 0, "rrc_users": 0, "ul_prb": 0, "dl_prb": 0, "voice_erl": 0}
        merged_data[key]["rrc_users"] = int(row.get("rrc_connmax") or 0)
    
    # 处理接通率数据
    for row in connect_data:
        key = (str(row.get("start_time", "")), row.get("network_type", ""))
        if key not in merged_data:
            merged_data[key] = {"traffic": 0, "connect_rate": 0, "rrc_users": 0, "ul_prb": 0, "dl_prb": 0, "voice_erl": 0}
        merged_data[key]["connect_rate"] = float(row.get("connect_rate") or 0)
    
    # 处理话务量数据
    for row in voice_data:
        key = (str(row.get("start_time", "")), row.get("network_type", ""))
        if key not in merged_data:
            merged_data[key] = {"traffic": 0, "connect_rate": 0, "rrc_users": 0, "ul_prb": 0, "dl_prb": 0, "voice_erl": 0}
        merged_data[key]["voice_erl"] = float(row.get("voice_erl") or 0)
    
    # 写入全网数据
    for (time_str, network), data in sorted(merged_data.items()):
        ws1.append([
            time_str,
            network,
            f"{data['connect_rate']:.2f}",
            f"{data['traffic']:.2f}",
            f"{data['ul_prb']:.2f}",
            f"{data['dl_prb']:.2f}",
            f"{data['voice_erl']:.2f}",
            data['rrc_users']
        ])
    
    # Sheet2: 地区数据（无线接通，流量，用户数）
    ws2 = wb.create_sheet("地区数据")
    
    # 地区数据列名
    region_headers = ["时间", "网络类型", "地区", "无线接通率(%)", "总流量(GB)", "上行PRB利用率(%)", "下行PRB利用率(%)", "话务量(Erl)", "RRC总用户数"]
    ws2.append(region_headers)
    
    # 处理地区数据
    # 按时间粒度查询地区数据
    # 1. 首先获取所有小区数据
    all_cells_data = []
    if "4G" in networks:
        table_4g = service.get_table_name("4G", granularity)
        sql_4g = f"""
            SELECT
                start_time,
                cellname,
                cgi,
                SUM(COALESCE("PDCP_UpOctUl",0) + COALESCE("PDCP_UpOctDl",0)) / 1000.0 / 1000.0 AS total_traffic,
                SUM("RRC_SuccConnEstab") * 100.0 / NULLIF(SUM("RRC_AttConnEstab"), 0) *
                SUM("ERAB_NbrSuccEstab") * 100.0 / NULLIF(SUM("ERAB_NbrAttEstab"), 0) / 100.0 AS wireless_connect_rate,
                SUM("RRU_PuschPrbAssn") * 100.0 / NULLIF(SUM("RRU_PuschPrbTot"), 0) AS ul_prb_util,
                SUM("RRU_PdschPrbAssn") * 100.0 / NULLIF(SUM("RRU_PdschPrbTot"), 0) AS dl_prb_util,
                SUM(COALESCE("ERAB_NbrMeanEstab_1", 0)) / 4.0 AS voice_erl,
                SUM("RRC_ConnMax") AS rrc_users
            FROM {table_4g}
            WHERE start_time BETWEEN %s AND %s
            GROUP BY start_time, cellname, cgi
        """
        cells_4g = service.pg.fetch_all(sql_4g, (start, end)) or []
        for cell in cells_4g:
            cell['network_type'] = '4G'
        all_cells_data.extend(cells_4g)
    
    if "5G" in networks:
        table_5g = service.get_table_name("5G", granularity)
        sql_5g = f"""
            SELECT
                start_time,
                userlabel AS cellname,
                "Ncgi" AS cgi,
                SUM(COALESCE("RLC_UpOctUl",0) + COALESCE("RLC_UpOctDl",0)) / 1000.0 / 1000.0 AS total_traffic,
                (SUM("RRC_SuccConnEstab") * 100.0 / NULLIF(SUM("RRC_AttConnEstab"), 0)) *
                (SUM("NGSIG_ConnEstabSucc") * 100.0 / NULLIF(SUM("NGSIG_ConnEstabAtt"), 0)) *
                (SUM("Flow_NbrSuccEstab") * 100.0 / NULLIF(SUM("Flow_NbrAttEstab"), 0)) / 100.0 / 100.0 AS wireless_connect_rate,
                SUM("RRU_PuschPrbAssn") * 100.0 / NULLIF(SUM("RRU_PuschPrbTot"), 0) AS ul_prb_util,
                SUM("RRU_PdschPrbAssn") * 100.0 / NULLIF(SUM("RRU_PdschPrbTot"), 0) AS dl_prb_util,
                SUM(COALESCE("Flow_NbrMeanEstab_5QI1", 0)) / 4.0 AS voice_erl,
                SUM("RRC_ConnMax") AS rrc_users
            FROM {table_5g}
            WHERE start_time BETWEEN %s AND %s
            GROUP BY start_time, userlabel, "Ncgi"
        """
        cells_5g = service.pg.fetch_all(sql_5g, (start, end)) or []
        for cell in cells_5g:
            cell['network_type'] = '5G'
        all_cells_data.extend(cells_5g)
    
    # 2. 按时间、网络类型和地区分组
    grouped_data = {}
    for cell in all_cells_data:
        time_str = str(cell.get("start_time", ""))
        network = cell.get("network_type", "")
        cellname = cell.get("cellname", "")
        cgi = cell.get("cgi", "")
        
        # 分类地区
        region = service.classify_region(cellname, network, cgi)
        
        key = (time_str, network, region)
        if key not in grouped_data:
            grouped_data[key] = {
                "total_traffic": 0,
                "total_connect_rate": 0,
                "total_ul_prb": 0,
                "total_dl_prb": 0,
                "total_voice_erl": 0,
                "total_rrc_users": 0,
                "count": 0
            }
        
        # 累加数据
        grouped_data[key]["total_traffic"] += float(cell.get("total_traffic", 0) or 0)
        connect_rate = float(cell.get("wireless_connect_rate", 0) or 0)
        if connect_rate > 0:
            grouped_data[key]["total_connect_rate"] += connect_rate
            grouped_data[key]["count"] += 1
        ul_prb = float(cell.get("ul_prb_util", 0) or 0)
        if ul_prb > 0:
            grouped_data[key]["total_ul_prb"] += ul_prb
        dl_prb = float(cell.get("dl_prb_util", 0) or 0)
        if dl_prb > 0:
            grouped_data[key]["total_dl_prb"] += dl_prb
        grouped_data[key]["total_voice_erl"] += float(cell.get("voice_erl", 0) or 0)
        grouped_data[key]["total_rrc_users"] += int(cell.get("rrc_users", 0) or 0)
    
    # 3. 写入地区数据
    for (time_str, network, region), data in sorted(grouped_data.items()):
        # 计算平均值
        avg_connect_rate = data["total_connect_rate"] / data["count"] if data["count"] > 0 else 0
        avg_ul_prb = data["total_ul_prb"] / data["count"] if data["count"] > 0 else 0
        avg_dl_prb = data["total_dl_prb"] / data["count"] if data["count"] > 0 else 0
        
        ws2.append([
            time_str,
            network,
            region,
            f"{avg_connect_rate:.2f}",
            f"{data['total_traffic']:.2f}",
            f"{avg_ul_prb:.2f}",
            f"{avg_dl_prb:.2f}",
            f"{data['total_voice_erl']:.2f}",
            data['total_rrc_users']
        ])
    
    # 如果没有数据，添加默认数据
    if not grouped_data:
        regions = ["江城区", "阳东县", "南区", "阳西县", "阳春市"]
        for region in regions:
            for network in networks:
                ws2.append([
                    start.strftime("%Y-%m-%d %H:%M:%S"),
                    network,
                    region,
                    "0.00",
                    "0.00",
                    "0.00",
                    "0.00",
                    "0.00",
                    "0"
                ])
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for ws in [ws1, ws2]:
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
    
    # 保存工作簿
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"仪表盘数据_{range_key}_{granularity}_{start.strftime('%Y%m%d_%H%M')}_{end.strftime('%Y%m%d_%H%M')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# 场景相关导出路由
@export_bp.route("/scenarios/download_template")
def download_cell_template():
    """下载小区导入模板"""
    # 使用通用Excel导出函数创建模板
    headers = ["cell_id", "cell_name", "cgi", "network_type"]
    sample_data = [
        {"cell_id": "123456-7", "cell_name": "示例小区1", "cgi": "460-00-123456-7", "network_type": "4G"},
        {"cell_id": "234567-8", "cell_name": "示例小区2", "cgi": "460-00-234567-8", "network_type": "5G"}
    ]
    
    column_widths = {
        "cell_id": 20,
        "cell_name": 20,
        "cgi": 25,
        "network_type": 15
    }
    
    output = create_export_workbook(
        sheet_name="小区导入模板",
        headers=headers,
        data=sample_data,
        column_widths=column_widths
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name="场景小区导入模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/scenarios/export_cells")
def export_scenario_cells():
    """导出场景下的小区列表"""
    scenario_service = current_app.config.get('scenario_service')
    
    if not scenario_service:
        flash("服务未初始化，无法导出数据", "danger")
        return redirect(url_for("scenarios"))
    
    scenario_id = int(request.args.get("scenario_id", 0))
    if not scenario_id:
        flash("请选择场景", "warning")
        return redirect(url_for("scenarios"))
    
    cells = scenario_service.list_cells(scenario_id)
    scenario_name = scenario_service.get_scenario_name(scenario_id) or f"场景_{scenario_id}"
    
    # 设置表头
    headers = ["cell_id", "cell_name", "cgi", "network_type"]
    
    # 列宽配置
    column_widths = {
        "cell_id": 20,
        "cell_name": 20,
        "cgi": 25,
        "network_type": 15
    }
    
    # 使用通用Excel导出函数
    output = create_export_workbook(
        sheet_name="小区列表",
        headers=headers,
        data=cells,
        column_widths=column_widths
    )
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{scenario_name}_小区列表.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )