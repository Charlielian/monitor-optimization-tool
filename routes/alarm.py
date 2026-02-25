"""
Alarm Routes Blueprint

This module contains all alarm-related routes:
- ZTE (中兴) alarm monitoring: 告警监控（中兴设备）
- Nokia (诺基亚) alarm monitoring: 告警监控（诺基亚设备）
- Alarm export routes: 告警导出

Requirements: 3.1, 3.2, 3.3, 3.4, 3.5
"""

import io
import logging
import time
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from auth import create_page_decorator, login_required, api_login_required
from utils.validators import sanitize_search_query


# Create the alarm blueprint with no URL prefix to maintain existing URL patterns
alarm_bp = Blueprint('alarm', __name__)


def get_alarm_context():
    """
    Get alarm-related context from app config.
    
    Returns:
        tuple: (alarm_service_zte, alarm_service_nokia, auth_enabled)
    """
    alarm_service_zte = current_app.config.get('alarm_service_zte')
    alarm_service_nokia = current_app.config.get('alarm_service_nokia')
    auth_enabled = current_app.config.get('auth_enabled', True)
    
    return alarm_service_zte, alarm_service_nokia, auth_enabled


def render_alarm_page(
    alarm_service,
    template_name: str,
    vendor_name: str,
    page_route: str
) -> str:
    """
    通用告警页面渲染函数
    
    Args:
        alarm_service: 告警服务实例
        template_name: 模板名称
        vendor_name: 厂商名称（用于日志）
        page_route: 页面路由名称（用于重定向）
    
    Returns:
        渲染后的HTML响应
    """
    from utils.time_parser import format_datetime_for_input
    
    route_start = time.time()
    
    if not alarm_service:
        flash(f"{vendor_name}告警服务未初始化，请检查MySQL配置", "danger")
        return redirect(url_for("dashboard"))
    
    # 获取tab参数
    active_tab = request.args.get("tab", "current")
    logging.info(f"📋 {vendor_name}告警页面请求 - Tab: {active_tab}")
    
    # 获取过滤参数
    param_start = time.time()
    current_ne_id_filter = sanitize_search_query(request.args.get("current_ne_id", ""))
    current_alarm_name_filter = sanitize_search_query(request.args.get("current_alarm_name", ""))
    hist_ne_id_filter = sanitize_search_query(request.args.get("hist_ne_id", ""))
    hist_alarm_name_filter = sanitize_search_query(request.args.get("hist_alarm_name", ""))
    logging.debug(f"  ├─ 参数解析: {(time.time() - param_start) * 1000:.2f}ms")

    # 获取告警统计（总是需要）
    stats_start = time.time()
    stats = alarm_service.get_alarm_statistics()
    stats_elapsed = (time.time() - stats_start) * 1000
    if stats_elapsed > 500:
        logging.warning(f"  ├─ ⚠️ {vendor_name}告警统计查询慢: {stats_elapsed:.2f}ms")
    else:
        logging.info(f"  ├─ {vendor_name}告警统计查询: {stats_elapsed:.2f}ms")

    # 根据活动标签页按需加载数据（优化性能）
    current_alarms = []
    historical_data = {"alarms": [], "total": 0, "page": 1, "pages": 0}

    if active_tab == "current":
        # 仅在当前告警标签页时加载当前告警
        current_start = time.time()
        current_alarms = alarm_service.get_current_alarms(
            ne_id_filter=current_ne_id_filter,
            alarm_name_filter=current_alarm_name_filter
        )
        current_elapsed = (time.time() - current_start) * 1000
        if current_elapsed > 1000:
            logging.warning(f"  ├─ ⚠️ {vendor_name}当前告警查询慢: {current_elapsed:.2f}ms (返回 {len(current_alarms)} 条)")
        else:
            logging.info(f"  ├─ {vendor_name}当前告警查询: {current_elapsed:.2f}ms (返回 {len(current_alarms)} 条)")
    elif active_tab == "historical":
        # 仅在历史告警标签页时加载历史告警
        # 获取历史告警查询参数
        start_time_str = request.args.get("start_time", "")
        end_time_str = request.args.get("end_time", "")
        page = int(request.args.get("page", 1))

        # 解析时间
        parse_start = time.time()
        start_time = None
        end_time = None
        if start_time_str:
            try:
                start_time = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M")
            except:
                pass
        if end_time_str:
            try:
                end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M")
            except:
                pass
        logging.debug(f"  ├─ 时间解析: {(time.time() - parse_start) * 1000:.2f}ms")

        # 获取历史告警（支持过滤）
        hist_start = time.time()
        historical_data = alarm_service.get_historical_alarms(
            start_time=start_time,
            end_time=end_time,
            page=page,
            page_size=20,
            ne_id_filter=hist_ne_id_filter,
            alarm_name_filter=hist_alarm_name_filter
        )
        hist_elapsed = (time.time() - hist_start) * 1000
        if hist_elapsed > 1000:
            logging.warning(f"  ├─ ⚠️ {vendor_name}历史告警查询慢: {hist_elapsed:.2f}ms (返回 {len(historical_data.get('alarms', []))} 条)")
        else:
            logging.info(f"  ├─ {vendor_name}历史告警查询: {hist_elapsed:.2f}ms (返回 {len(historical_data.get('alarms', []))} 条)")

    # 格式化时间用于表单显示
    format_start = time.time()
    start_time_str = request.args.get("start_time", "")
    end_time_str = request.args.get("end_time", "")
    if not start_time_str and historical_data.get('start_time'):
        start_time_str = format_datetime_for_input(historical_data['start_time'])
    if not end_time_str and historical_data.get('end_time'):
        end_time_str = format_datetime_for_input(historical_data['end_time'])
    logging.debug(f"  ├─ 时间格式化: {(time.time() - format_start) * 1000:.2f}ms")
    
    # 渲染模板
    render_start = time.time()
    response = render_template(
        template_name,
        active_tab=active_tab,
        stats=stats,
        current_alarms=current_alarms,
        historical_data=historical_data,
        start_time=start_time_str,
        end_time=end_time_str,
        current_ne_id_filter=current_ne_id_filter,
        current_alarm_name_filter=current_alarm_name_filter,
        hist_ne_id_filter=hist_ne_id_filter,
        hist_alarm_name_filter=hist_alarm_name_filter,
    )
    render_elapsed = (time.time() - render_start) * 1000
    if render_elapsed > 500:
        logging.warning(f"  ├─ ⚠️ 模板渲染慢: {render_elapsed:.2f}ms")
    else:
        logging.info(f"  ├─ 模板渲染: {render_elapsed:.2f}ms")
    
    total_elapsed = (time.time() - route_start) * 1000
    if total_elapsed > 3000:
        logging.error(f"  └─ 🔴 {vendor_name}告警页面总耗时: {total_elapsed:.2f}ms (超过3秒)")
    elif total_elapsed > 1000:
        logging.warning(f"  └─ 🟡 {vendor_name}告警页面总耗时: {total_elapsed:.2f}ms (超过1秒)")
    else:
        logging.info(f"  └─ ✅ {vendor_name}告警页面总耗时: {total_elapsed:.2f}ms")
    
    return response


def export_current_alarms_common(
    alarm_service,
    vendor_name: str,
    redirect_route: str
):
    """
    通用当前告警导出函数
    
    Args:
        alarm_service: 告警服务实例
        vendor_name: 厂商名称
        redirect_route: 重定向路由名称
    
    Returns:
        Excel文件响应或重定向
    """
    if not alarm_service:
        flash(f"{vendor_name}告警服务未初始化", "danger")
        return redirect(url_for(redirect_route))

    try:
        from utils.excel_helper import create_alarm_export

        # 获取过滤参数
        ne_id_filter = request.args.get("current_ne_id", "")
        alarm_name_filter = request.args.get("current_alarm_name", "")

        # 使用相同的过滤条件获取告警数据
        alarms = alarm_service.get_current_alarms(
            ne_id_filter=ne_id_filter,
            alarm_name_filter=alarm_name_filter
        )

        # 定义表头
        headers = [
            "告警时间", "告警级别", "告警名称", "告警类型",
            "网元", "网元名称", "网元ID", "网元类型",
            "站点名称", "告警对象", "位置", "附加信息",
            "确认状态", "告警原因"
        ]

        # 定义列宽
        column_widths = {
            "告警时间": 20, "告警级别": 12, "告警名称": 30, "告警类型": 15,
            "网元": 20, "网元名称": 15, "网元ID": 20, "网元类型": 12,
            "站点名称": 25, "告警对象": 25, "位置": 25, "附加信息": 30,
            "确认状态": 12, "告警原因": 40
        }

        # 定义数据映射函数
        def map_alarm_data(alarm):
            return [
                str(alarm.get("occur_time", "")),
                alarm.get("alarm_level", ""),
                alarm.get("alarm_code_name", ""),
                alarm.get("alarm_type", ""),
                alarm.get("网元", ""),
                "-",
                alarm.get("ne_id", ""),
                alarm.get("ne_type", ""),
                alarm.get("site_name", ""),
                alarm.get("alarm_object_name", ""),
                alarm.get("location", ""),
                alarm.get("additional_info", ""),
                alarm.get("ack_status", ""),
                alarm.get("alarm_reason", ""),
            ]

        # 使用通用函数创建Excel
        sheet_name = f"{vendor_name}当前告警" if vendor_name != "中兴" else "当前告警"
        output = create_alarm_export(
            sheet_name=sheet_name,
            headers=headers,
            data=alarms,
            column_widths=column_widths,
            data_mapper=map_alarm_data,
            filename_prefix=f"{vendor_name}当前告警" if vendor_name != "中兴" else "当前告警"
        )

        filename = f"{vendor_name}当前告警_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx" if vendor_name != "中兴" else f"当前告警_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logging.error(f"导出{vendor_name}当前告警失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for(redirect_route))


def export_historical_alarms_common(
    alarm_service,
    vendor_name: str,
    redirect_route: str
):
    """
    通用历史告警导出函数
    
    Args:
        alarm_service: 告警服务实例
        vendor_name: 厂商名称
        redirect_route: 重定向路由名称
    
    Returns:
        Excel文件响应或重定向
    """
    if not alarm_service:
        flash(f"{vendor_name}告警服务未初始化", "danger")
        return redirect(url_for(redirect_route))
    
    try:
        # 获取时间参数和过滤参数
        start_time_str = request.args.get("start_time", "")
        end_time_str = request.args.get("end_time", "")
        ne_id_filter = request.args.get("hist_ne_id", "")
        alarm_name_filter = request.args.get("hist_alarm_name", "")
        
        start_time = None
        end_time = None
        if start_time_str:
            try:
                start_time = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M")
            except:
                pass
        if end_time_str:
            try:
                end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M")
            except:
                pass
        
        # 获取所有历史告警（不分页，使用相同的过滤条件）
        historical_data = alarm_service.get_historical_alarms(
            start_time=start_time,
            end_time=end_time,
            page=1,
            page_size=10000,  # 导出时获取更多数据
            ne_id_filter=ne_id_filter,
            alarm_name_filter=alarm_name_filter
        )
        alarms = historical_data.get('alarms', [])
        
        wb = Workbook()
        ws = wb.active
        sheet_name = f"{vendor_name}历史告警" if vendor_name != "中兴" else "历史告警"
        ws.title = sheet_name
        
        # 中文表头（与页面显示一致）
        headers = [
            "告警时间", "告警级别", "告警名称", "告警类型", 
            "网元", "网元名称", "网元ID", "网元类型", 
            "站点名称", "告警对象", "位置", "附加信息",
            "确认状态", "告警原因"
        ]
        ws.append(headers)
        
        # 添加数据
        for alarm in alarms:
            ws.append([
                str(alarm.get("occur_time", "")),
                alarm.get("alarm_level", ""),
                alarm.get("alarm_code_name", ""),
                alarm.get("alarm_type", ""),
                alarm.get("网元", ""),
                "-",  # 网元名称字段不存在
                alarm.get("ne_id", ""),
                alarm.get("ne_type", ""),
                alarm.get("site_name", ""),
                alarm.get("alarm_object_name", ""),
                alarm.get("location", ""),
                alarm.get("additional_info", ""),
                alarm.get("ack_status", ""),
                alarm.get("alarm_reason", ""),
            ])
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20  # 告警时间
        ws.column_dimensions['B'].width = 12  # 告警级别
        ws.column_dimensions['C'].width = 30  # 告警名称
        ws.column_dimensions['D'].width = 15  # 告警类型
        ws.column_dimensions['E'].width = 20  # 网元
        ws.column_dimensions['F'].width = 15  # 网元名称
        ws.column_dimensions['G'].width = 20  # 网元ID
        ws.column_dimensions['H'].width = 12  # 网元类型
        ws.column_dimensions['I'].width = 25  # 站点名称
        ws.column_dimensions['J'].width = 25  # 告警对象
        ws.column_dimensions['K'].width = 25  # 位置
        ws.column_dimensions['L'].width = 30  # 附加信息
        ws.column_dimensions['M'].width = 12  # 确认状态
        ws.column_dimensions['N'].width = 40  # 告警原因
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{vendor_name}历史告警_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx" if vendor_name != "中兴" else f"历史告警_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logging.error(f"导出{vendor_name}历史告警失败: {e}")
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for(redirect_route))


# ==================== 中兴设备告警路由 ====================

@alarm_bp.route("/alarm")
@login_required
def alarm():
    """中兴设备告警监控查询页面"""
    alarm_service_zte, _, auth_enabled = get_alarm_context()
    
    return render_alarm_page(
        alarm_service=alarm_service_zte,
        template_name="alarm.html",
        vendor_name="中兴",
        page_route="alarm.alarm"
    )


@alarm_bp.route("/export/current_alarms.xlsx")
def export_current_alarms():
    """导出中兴当前告警为Excel"""
    alarm_service_zte, _, auth_enabled = get_alarm_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    return export_current_alarms_common(
        alarm_service=alarm_service_zte,
        vendor_name="中兴",
        redirect_route="alarm.alarm"
    )


@alarm_bp.route("/export/historical_alarms.xlsx")
def export_historical_alarms():
    """导出中兴历史告警为Excel"""
    alarm_service_zte, _, auth_enabled = get_alarm_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    return export_historical_alarms_common(
        alarm_service=alarm_service_zte,
        vendor_name="中兴",
        redirect_route="alarm.alarm"
    )


# ==================== 诺基亚设备告警路由 ====================

@alarm_bp.route("/alarm_nokia")
@login_required
def alarm_nokia():
    """诺基亚设备告警监控查询页面"""
    _, alarm_service_nokia, auth_enabled = get_alarm_context()
    
    return render_alarm_page(
        alarm_service=alarm_service_nokia,
        template_name="alarm_nokia.html",
        vendor_name="诺基亚",
        page_route="alarm.alarm_nokia"
    )


@alarm_bp.route("/export/current_alarms_nokia.xlsx")
def export_current_alarms_nokia():
    """导出诺基亚当前告警为Excel"""
    _, alarm_service_nokia, auth_enabled = get_alarm_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    return export_current_alarms_common(
        alarm_service=alarm_service_nokia,
        vendor_name="诺基亚",
        redirect_route="alarm.alarm_nokia"
    )


@alarm_bp.route("/export/historical_alarms_nokia.xlsx")
def export_historical_alarms_nokia():
    """导出诺基亚历史告警为Excel"""
    _, alarm_service_nokia, auth_enabled = get_alarm_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    return export_historical_alarms_common(
        alarm_service=alarm_service_nokia,
        vendor_name="诺基亚",
        redirect_route="alarm.alarm_nokia"
    )


# ==================== 自动完成API ====================

@alarm_bp.route("/api/autocomplete/ne_site")
@api_login_required
def api_autocomplete_ne_site():
    """
    获取网元ID和站点名称的自动完成数据（中兴设备）
    
    Returns:
        JSON: {ne_ids: [...], site_names: [...]}
    """
    from flask import jsonify
    
    alarm_service_zte, _, _ = get_alarm_context()
    
    if not alarm_service_zte or not alarm_service_zte.mysql:
        return jsonify({"ne_ids": [], "site_names": []})
    
    try:
        # 获取最近的网元ID和站点名称（去重）
        query = f"""
        SELECT DISTINCT ne_id, site_name 
        FROM {alarm_service_zte.current_table}
        WHERE ne_id IS NOT NULL AND ne_id != ''
        ORDER BY occur_time DESC
        LIMIT 500
        """
        
        results = alarm_service_zte.mysql.fetch_all(query)
        
        ne_ids = set()
        site_names = set()
        
        for row in results:
            if row.get('ne_id'):
                ne_ids.add(row['ne_id'])
            if row.get('site_name'):
                site_names.add(row['site_name'])
        
        return jsonify({
            "ne_ids": sorted(list(ne_ids)),
            "site_names": sorted(list(site_names))
        })
    except Exception as e:
        logging.error(f"获取网元/站点自动完成数据失败: {e}")
        return jsonify({"ne_ids": [], "site_names": []})


@alarm_bp.route("/api/autocomplete/alarm_names")
@api_login_required
def api_autocomplete_alarm_names():
    """
    获取告警名称的自动完成数据（中兴设备）
    
    Returns:
        JSON: {alarm_names: [...]}
    """
    from flask import jsonify
    
    alarm_service_zte, _, _ = get_alarm_context()
    
    if not alarm_service_zte or not alarm_service_zte.mysql:
        return jsonify({"alarm_names": []})
    
    try:
        # 获取最近的告警名称（去重）
        query = f"""
        SELECT DISTINCT alarm_code_name 
        FROM {alarm_service_zte.current_table}
        WHERE alarm_code_name IS NOT NULL AND alarm_code_name != ''
        ORDER BY occur_time DESC
        LIMIT 200
        """
        
        results = alarm_service_zte.mysql.fetch_all(query)
        alarm_names = sorted(list(set(row['alarm_code_name'] for row in results if row.get('alarm_code_name'))))
        
        return jsonify({"alarm_names": alarm_names})
    except Exception as e:
        logging.error(f"获取告警名称自动完成数据失败: {e}")
        return jsonify({"alarm_names": []})


@alarm_bp.route("/api/autocomplete/ne_site_nokia")
@api_login_required
def api_autocomplete_ne_site_nokia():
    """
    获取网元ID和站点名称的自动完成数据（诺基亚设备）
    
    Returns:
        JSON: {ne_ids: [...], site_names: [...]}
    """
    from flask import jsonify
    
    _, alarm_service_nokia, _ = get_alarm_context()
    
    if not alarm_service_nokia or not alarm_service_nokia.mysql:
        return jsonify({"ne_ids": [], "site_names": []})
    
    try:
        # 获取最近的网元ID和站点名称（去重）
        query = f"""
        SELECT DISTINCT ne_id, site_name 
        FROM {alarm_service_nokia.current_table}
        WHERE ne_id IS NOT NULL AND ne_id != ''
        ORDER BY occur_time DESC
        LIMIT 500
        """
        
        results = alarm_service_nokia.mysql.fetch_all(query)
        
        ne_ids = set()
        site_names = set()
        
        for row in results:
            if row.get('ne_id'):
                ne_ids.add(row['ne_id'])
            if row.get('site_name'):
                site_names.add(row['site_name'])
        
        return jsonify({
            "ne_ids": sorted(list(ne_ids)),
            "site_names": sorted(list(site_names))
        })
    except Exception as e:
        logging.error(f"获取诺基亚网元/站点自动完成数据失败: {e}")
        return jsonify({"ne_ids": [], "site_names": []})


@alarm_bp.route("/api/autocomplete/alarm_names_nokia")
@api_login_required
def api_autocomplete_alarm_names_nokia():
    """
    获取告警名称的自动完成数据（诺基亚设备）
    
    Returns:
        JSON: {alarm_names: [...]}
    """
    from flask import jsonify
    
    _, alarm_service_nokia, _ = get_alarm_context()
    
    if not alarm_service_nokia or not alarm_service_nokia.mysql:
        return jsonify({"alarm_names": []})
    
    try:
        # 获取最近的告警名称（去重）
        query = f"""
        SELECT DISTINCT alarm_code_name 
        FROM {alarm_service_nokia.current_table}
        WHERE alarm_code_name IS NOT NULL AND alarm_code_name != ''
        ORDER BY occur_time DESC
        LIMIT 200
        """
        
        results = alarm_service_nokia.mysql.fetch_all(query)
        alarm_names = sorted(list(set(row['alarm_code_name'] for row in results if row.get('alarm_code_name'))))
        
        return jsonify({"alarm_names": alarm_names})
    except Exception as e:
        logging.error(f"获取诺基亚告警名称自动完成数据失败: {e}")
        return jsonify({"alarm_names": []})
