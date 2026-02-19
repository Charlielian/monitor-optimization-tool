"""
Grid Monitoring Routes Blueprint

This module contains all grid monitoring routes:
- Grid list page: 网格监控列表
- Grid detail page: 网格详情
- Grid export routes: 网格数据导出
- Grid autocomplete API: 网格自动联想

Requirements: 4.1, 4.2, 4.3, 4.4, 4.5
"""

import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from flask import (
    Blueprint,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook

from auth import create_page_decorator, login_required, admin_required, api_login_required
from services.cache import cache_5m, cache_30m
from services.grid_health_check import GridHealthCheckService
from services.hsr_health_check import HSRHealthCheckService
from utils.validators import sanitize_search_query
from flask import session


# Create the grid blueprint with no URL prefix to maintain existing URL patterns
grid_bp = Blueprint('grid', __name__)


def get_grid_context():
    """
    Get grid-related context from app config.
    
    Returns:
        tuple: (grid_service, health_check_service, auth_enabled)
    """
    grid_service = current_app.config.get('grid_service')
    auth_enabled = current_app.config.get('auth_enabled', True)
    
    # 初始化体检服务
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    health_check_service = GridHealthCheckService(mysql_client, pg_client) if mysql_client else None
    
    return grid_service, health_check_service, auth_enabled


def format_traffic(gb_value: float) -> str:
    """
    智能转换流量单位 (GB/TB)
    
    Args:
        gb_value: 流量值（GB）
    
    Returns:
        格式化后的流量字符串
    """
    if gb_value >= 1000:
        return f"{gb_value / 1000:.2f} TB"
    else:
        return f"{gb_value:.2f} GB"


# ==================== 网格监控路由 ====================

@grid_bp.route("/grid")
@login_required
def grid():
    """网格监控列表页面"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    try:
        # 获取搜索参数
        search = sanitize_search_query(request.args.get("search", "").strip())
        # 获取对比模式参数（默认为天级对比）
        comparison_mode = request.args.get("comparison_mode", "daily")
        
        # 初始化变量
        grids = []
        dashboard_stats = None
        
        # 检查服务是否可用
        if not grid_service:
            logging.warning("网格监控服务未初始化")
            flash("网格监控服务未初始化，请检查数据库配置", "warning")
            # 即使服务不可用，也显示主界面（但没有数据）
            return render_template(
                "grid.html",
                grids=grids,
                search=search,
                dashboard_stats=dashboard_stats,
                comparison_mode=comparison_mode
            )
        
        # 获取网格列表
        try:
            grids = grid_service.get_grid_list(search=search if search else None)
            logging.info(f"获取网格列表成功，共 {len(grids)} 个网格")
        except Exception as e:
            logging.error(f"获取网格列表失败: {e}", exc_info=True)
            flash("获取网格列表失败，请检查数据库连接", "warning")
        
        # 获取仪表盘统计数据（始终显示，使用5分钟缓存）
        try:
            # 获取阈值参数
            prb_threshold_4g = float(request.args.get("prb_4g", 50.0))
            prb_threshold_5g = float(request.args.get("prb_5g", 50.0))
            
            logging.info(f"开始获取仪表盘统计数据，阈值: 4G={prb_threshold_4g}%, 5G={prb_threshold_5g}%, 对比模式: {comparison_mode}")
            
            # 使用缓存（5分钟）
            cache_key = f"grid_dashboard:{prb_threshold_4g}:{prb_threshold_5g}:{comparison_mode}"
            dashboard_stats = cache_5m.get(
                cache_key,
                lambda: grid_service.get_dashboard_stats(
                    prb_threshold_4g=prb_threshold_4g,
                    prb_threshold_5g=prb_threshold_5g,
                    comparison_mode=comparison_mode
                )
            )
            logging.info(f"仪表盘统计数据获取成功")
            
            # 为网格列表添加问题标签
            if dashboard_stats and grids:
                # 构建问题网格的映射
                grid_issues = {}
                
                # 流量劣化网格
                if dashboard_stats.get('traffic_degraded_grids'):
                    for grid_data in dashboard_stats['traffic_degraded_grids']:
                        grid_id = grid_data['grid_id']
                        if grid_id not in grid_issues:
                            grid_issues[grid_id] = []
                        grid_issues[grid_id].append({
                            'type': 'traffic_degraded',
                            'label': '流量劣化',
                            'class': 'warning',
                            'detail': f"{grid_data['change_rate']:.1f}%"
                        })
                
                # 无流量小区增加网格
                if dashboard_stats.get('no_traffic_increased_grids'):
                    for grid_data in dashboard_stats['no_traffic_increased_grids']:
                        grid_id = grid_data['grid_id']
                        if grid_id not in grid_issues:
                            grid_issues[grid_id] = []
                        grid_issues[grid_id].append({
                            'type': 'no_traffic_increased',
                            'label': '无流量小区增加',
                            'class': 'info',
                            'detail': f"+{grid_data['increase_rate']:.1f}%"
                        })
                
                # 高负荷小区网格
                if dashboard_stats.get('high_load_grids'):
                    for grid_data in dashboard_stats['high_load_grids']:
                        grid_id = grid_data['grid_id']
                        if grid_id not in grid_issues:
                            grid_issues[grid_id] = []
                        grid_issues[grid_id].append({
                            'type': 'high_load',
                            'label': '高负荷小区',
                            'class': 'success',
                            'detail': f"{grid_data['total_high_load']}个"
                        })
                
                # 故障网格
                if dashboard_stats.get('fault_grids_detail'):
                    for grid_data in dashboard_stats['fault_grids_detail']:
                        grid_id = grid_data['grid_id']
                        if grid_id not in grid_issues:
                            grid_issues[grid_id] = []
                        grid_issues[grid_id].append({
                            'type': 'fault',
                            'label': '故障',
                            'class': 'primary',
                            'detail': f"{grid_data['fault_count']}个"
                        })
                
                # 为每个网格添加问题标签
                for grid in grids:
                    grid['issues'] = grid_issues.get(grid['grid_id'], [])
            
        except Exception as e:
            logging.error(f"获取仪表盘统计数据失败: {e}", exc_info=True)
            dashboard_stats = None
        
        return render_template(
            "grid.html",
            grids=grids,
            search=search,
            dashboard_stats=dashboard_stats,
            comparison_mode=comparison_mode
        )
    except Exception as e:
        logging.error(f"网格监控页面渲染失败: {e}", exc_info=True)
        return f"Internal Server Error: {str(e)}", 500


@grid_bp.route("/api/grid/autocomplete")
@api_login_required
def grid_autocomplete():
    """网格自动联想API"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    try:
        # 获取查询参数
        query = request.args.get("q", "").strip()
        
        if not query:
            logging.debug(f"网格自动联想: 查询参数为空")
            return jsonify([])
        
        if not grid_service:
            logging.warning(f"网格自动联想: grid_service未初始化")
            return jsonify([])
        
        # 获取所有网格列表
        all_grids = grid_service.get_grid_list()
        logging.debug(f"网格自动联想: 查询'{query}'，总网格数={len(all_grids)}")
        
        # 过滤匹配的网格（网格ID或网格名称包含查询字符串）
        matched_grids = []
        for grid in all_grids:
            grid_id = grid.get('grid_id', '')
            grid_name = grid.get('grid_name', '')
            
            if query.lower() in grid_id.lower() or query.lower() in str(grid_name).lower():
                matched_grids.append({
                    'id': grid_id,
                    'name': grid_name or grid_id,
                    'label': f"{grid_id} - {grid_name}" if grid_name else grid_id,
                    'cell_4g_count': grid.get('cell_4g_count', 0),
                    'cell_5g_count': grid.get('cell_5g_count', 0),
                })
            
            # 限制返回数量
            if len(matched_grids) >= 10:
                break
        
        logging.info(f"网格自动联想: 查询'{query}'，匹配{len(matched_grids)}个网格")
        return jsonify(matched_grids)
        
    except Exception as e:
        logging.error(f"网格自动联想失败: {e}", exc_info=True)
        return jsonify([])


@grid_bp.route("/grid/<grid_id>")
@login_required
def grid_detail(grid_id):
    """网格详情页面"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 检查服务是否可用
    if not grid_service:
        flash("网格监控服务未初始化，请检查数据库配置", "warning")
        return redirect(url_for("grid.grid"))


# ==================== 高铁小区健康检查路由 ====================

@grid_bp.route("/hsr/health_check")
@login_required
def hsr_health_check():
    """高铁小区健康检查页面"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 初始化高铁体检服务
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    hsr_health_service = HSRHealthCheckService(mysql_client, pg_client) if mysql_client else None
    
    # 检查服务是否可用
    if not hsr_health_service:
        flash("高铁体检服务未初始化，请检查数据库配置", "warning")
        return redirect(url_for("grid.grid"))
    
    try:
        # 执行体检
        health_result = hsr_health_service.check_hsr_health()

        if 'error' in health_result:
            flash(f"体检失败: {health_result['error']}", "danger")
            return redirect(url_for("grid.grid"))

        # 缓存体检结果，供导出功能使用（30分钟有效期）
        # 使用用户名作为缓存键的一部分，确保每个用户有独立的缓存
        username = session.get('username', 'anonymous')
        cache_key = f"hsr_health_check:{username}"
        cache_30m.set(cache_key, health_result)
        logging.info(f"已缓存高铁体检结果（30分钟），缓存键: {cache_key}")

        # 按状态分组
        healthy_cells = [c for c in health_result['cells'] if c['status'] == 'healthy']
        unhealthy_cells = [c for c in health_result['cells'] if c['status'] == 'unhealthy']
        
        # 按原因分组不健康小区
        unhealthy_by_reason = {}
        for cell in unhealthy_cells:
            reason = cell['reason']
            if reason not in unhealthy_by_reason:
                unhealthy_by_reason[reason] = []
            unhealthy_by_reason[reason].append(cell)
        
        # 获取所有高铁线路
        hsr_lines = hsr_health_service.get_hsr_lines()
        
        return render_template(
            "hsr_health_check.html",
            check_time=health_result['check_time'],
            total_cells=health_result['total_cells'],
            healthy_cells_count=health_result['healthy_cells'],
            unhealthy_cells_count=health_result['unhealthy_cells'],
            healthy_rate=health_result['healthy_rate'],
            healthy_cells=healthy_cells,
            unhealthy_cells=unhealthy_cells,
            unhealthy_by_reason=unhealthy_by_reason,
            hsr_lines=hsr_lines
        )
        
    except Exception as e:
        logging.error(f"高铁体检页面渲染失败: {e}", exc_info=True)
        flash(f"体检失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))


@grid_bp.route("/hsr/health_check/<line_name>")
@login_required
def hsr_line_health_check(line_name):
    """特定高铁线路健康检查页面"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 初始化高铁体检服务
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    hsr_health_service = HSRHealthCheckService(mysql_client, pg_client) if mysql_client else None
    
    # 检查服务是否可用
    if not hsr_health_service:
        flash("高铁体检服务未初始化，请检查数据库配置", "warning")
        return redirect(url_for("grid.hsr_health_check"))
    
    try:
        # 执行体检
        health_result = hsr_health_service.check_hsr_line_health(line_name)
        
        if 'error' in health_result:
            flash(f"体检失败: {health_result['error']}", "danger")
            return redirect(url_for("grid.hsr_health_check"))
        
        # 按状态分组
        healthy_cells = [c for c in health_result['cells'] if c['status'] == 'healthy']
        unhealthy_cells = [c for c in health_result['cells'] if c['status'] == 'unhealthy']
        
        # 按原因分组不健康小区
        unhealthy_by_reason = {}
        for cell in unhealthy_cells:
            reason = cell['reason']
            if reason not in unhealthy_by_reason:
                unhealthy_by_reason[reason] = []
            unhealthy_by_reason[reason].append(cell)
        
        # 获取所有高铁线路
        hsr_lines = hsr_health_service.get_hsr_lines()
        
        return render_template(
            "hsr_health_check.html",
            line_name=line_name,
            check_time=health_result['check_time'],
            total_cells=health_result['total_cells'],
            healthy_cells_count=health_result['healthy_cells'],
            unhealthy_cells_count=health_result['unhealthy_cells'],
            healthy_rate=health_result['healthy_rate'],
            healthy_cells=healthy_cells,
            unhealthy_cells=unhealthy_cells,
            unhealthy_by_reason=unhealthy_by_reason,
            hsr_lines=hsr_lines
        )
        
    except Exception as e:
        logging.error(f"高铁线路体检页面渲染失败: {e}", exc_info=True)
        flash(f"体检失败: {str(e)}", "danger")
        return redirect(url_for("grid.hsr_health_check"))


@grid_bp.route("/hsr/health_check/export")
@login_required
def export_hsr_health_check():
    """导出高铁小区健康检查结果"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 初始化高铁体检服务
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    hsr_health_service = HSRHealthCheckService(mysql_client, pg_client) if mysql_client else None
    
    # 检查服务是否可用
    if not hsr_health_service:
        flash("高铁体检服务未初始化，请检查数据库配置", "warning")
        return redirect(url_for("grid.hsr_health_check"))
    
    try:
        # 使用缓存获取体检结果（30分钟有效期）
        # 如果缓存未命中，会自动调用loader函数执行体检
        username = session.get('username', 'anonymous')
        cache_key = f"hsr_health_check:{username}"

        def load_health_check():
            """加载高铁体检数据"""
            logging.info(f"缓存未命中，执行新的高铁体检，缓存键: {cache_key}")
            result = hsr_health_service.check_hsr_health()
            if 'error' in result:
                raise Exception(result['error'])
            return result

        # 使用缓存的get方法，自动处理缓存命中和未命中的情况
        health_result = cache_30m.get(cache_key, load_health_check)

        if 'error' in health_result:
            flash(f"体检失败: {health_result['error']}", "danger")
            return redirect(url_for("grid.hsr_health_check"))
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # ===== Sheet 1: 小区明细 =====
        ws1 = wb.active
        ws1.title = "高铁小区健康检查"
        
        # 设置表头
        headers1 = [
            '线路名称', '站点名称', 'Transmitting_Point_Name', '区域', '站点类型', 'BBU名称', '小区名称',
            'CGI', '制式', '健康状态', '原因', '有性能数据',
            '前一日流量(GB)', '当日累计流量(GB)', '前一日忙时流量(GB)', '当日忙时流量(GB)',
            '前一日忙时利用率(%)', '当日忙时利用率(%)',
            '有告警', '告警数量', '告警明细', 'RRU ID', 'RRU ID Key',
            'CP ID', 'CP ID Key', 'RRU类型', '检查时间'
        ]
        ws1.append(headers1)
        
        # 填充数据（先不健康后健康）
        cells_sorted = sorted(
            health_result['cells'], 
            key=lambda x: (0 if x.get('status') == 'unhealthy' else 1, x.get('line_name', ''), x.get('site_name', ''), x.get('cgi', ''))
        )
        
        for cell in cells_sorted:
            # 格式化告警明细
            alarm_details = []
            for alarm in cell.get('alarm_details', []):
                alarm_str = f"{alarm.get('vendor', '')}: {alarm.get('alarm_name', '')} ({alarm.get('alarm_level', '')})"
                alarm_details.append(alarm_str)
            alarm_details_str = '; '.join(alarm_details) if alarm_details else '无'
            
            ws1.append([
                                cell.get('line_name', ''),
                                cell.get('site_name', ''),
                                cell.get('site_name', ''),  # Transmitting_Point_Name (same as site_name)
                                cell.get('area', ''),
                                cell.get('site_type', ''),
                                cell.get('bbu_name', ''),
                                cell.get('celname', ''),
                                cell.get('cgi', ''),
                                cell.get('network_type', ''),
                                '健康' if cell.get('status') == 'healthy' else '不健康',
                                cell.get('reason', ''),
                                '是' if cell.get('has_performance') else '否',
                                cell.get('yesterday_traffic', 0.0),
                                cell.get('today_traffic', 0.0),
                                cell.get('yesterday_busy_hour_traffic', 0.0),
                                cell.get('today_busy_hour_traffic', 0.0),
                                cell.get('yesterday_busy_hour_util', 0.0),
                                cell.get('today_busy_hour_util', 0.0),
                                '是' if cell.get('has_alarm') else '否',
                                cell.get('alarm_count', 0),
                                alarm_details_str,
                                cell.get('rru_id', ''),
                                cell.get('rru_id_key', ''),
                                cell.get('cpId', ''),
                                cell.get('cpId_key', ''),
                                cell.get('rru_type', ''),
                                cell.get('check_time', '').strftime('%Y-%m-%d %H:%M:%S') if cell.get('check_time') else ''
                            ])
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 20  # 线路名称
        ws1.column_dimensions['B'].width = 30  # 站点名称
        ws1.column_dimensions['C'].width = 30  # Transmitting_Point_Name
        ws1.column_dimensions['D'].width = 10  # 区域
        ws1.column_dimensions['E'].width = 10  # 站点类型
        ws1.column_dimensions['F'].width = 30  # BBU名称
        ws1.column_dimensions['G'].width = 30  # 小区名称
        ws1.column_dimensions['H'].width = 25  # CGI
        ws1.column_dimensions['I'].width = 8   # 制式
        ws1.column_dimensions['J'].width = 10  # 健康状态
        ws1.column_dimensions['K'].width = 15  # 原因
        ws1.column_dimensions['L'].width = 10  # 有性能数据
        ws1.column_dimensions['M'].width = 15  # 前一日流量(GB)
        ws1.column_dimensions['N'].width = 15  # 当日累计流量(GB)
        ws1.column_dimensions['O'].width = 15  # 前一日忙时流量(GB)
        ws1.column_dimensions['P'].width = 15  # 当日忙时流量(GB)
        ws1.column_dimensions['Q'].width = 15  # 前一日忙时利用率(%)
        ws1.column_dimensions['R'].width = 15  # 当日忙时利用率(%)
        ws1.column_dimensions['S'].width = 10  # 有告警
        ws1.column_dimensions['T'].width = 10  # 告警数量
        ws1.column_dimensions['U'].width = 50  # 告警明细
        ws1.column_dimensions['V'].width = 12  # RRU ID
        ws1.column_dimensions['W'].width = 20  # RRU ID Key
        ws1.column_dimensions['X'].width = 10  # CP ID
        ws1.column_dimensions['Y'].width = 20  # CP ID Key
        ws1.column_dimensions['Z'].width = 20  # RRU类型
        ws1.column_dimensions['AA'].width = 20  # 检查时间
        
        # ===== Sheet 2: 汇总页（按发射点）=====
        ws2 = wb.create_sheet(title="汇总页")
        
        # 设置汇总页表头
        summary_headers = [
            '线路名称', 'Transmitting_Point_Name', '4G 故障', '5G 故障'
        ]
        ws2.append(summary_headers)
        
        # 按站点名称分组
        sites = {}
        for cell in health_result['cells']:
            site_name = cell.get('site_name', '')
            line_name = cell.get('line_name', '')
            
            if site_name not in sites:
                sites[site_name] = {
                    'line_name': line_name,
                    '4G': set(),
                    '5G': set()
                }
            
            network_type = cell.get('network_type', '')
            if network_type in ['4G', '5G'] and cell.get('has_alarm', False):
                rru_id = cell.get('rru_id', '')
                alarm_details = []
                for alarm in cell.get('alarm_details', []):
                    alarm_str = f"{alarm.get('vendor', '')}: {alarm.get('alarm_name', '')} ({alarm.get('alarm_level', '')})"
                    alarm_details.append(alarm_str)
                
                if alarm_details:
                    fault_str = f"故障 {rru_id} + {' ; '.join(alarm_details)}"
                    sites[site_name][network_type].add(fault_str)
        
        # 填充汇总数据
        for site_name, faults in sorted(sites.items()):
            # 合并4G故障（转换为列表并排序以保持一致性）
            fault_4g = '; '.join(sorted(faults['4G'])) if faults['4G'] else '无故障'
            # 合并5G故障（转换为列表并排序以保持一致性）
            fault_5g = '; '.join(sorted(faults['5G'])) if faults['5G'] else '无故障'
            
            ws2.append([
                faults.get('line_name', ''),
                site_name,
                fault_4g,
                fault_5g
            ])
        
        # 设置汇总页列宽
        ws2.column_dimensions['A'].width = 20  # 线路名称
        ws2.column_dimensions['B'].width = 30  # Transmitting_Point_Name
        ws2.column_dimensions['C'].width = 100  # 4G 故障
        ws2.column_dimensions['D'].width = 100  # 5G 故障
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"高铁小区健康检查_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出高铁小区健康检查数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.hsr_health_check"))


@grid_bp.route("/hsr/lines")
@login_required
def hsr_lines():
    """获取所有高铁线路的API"""
    # 初始化高铁体检服务
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    hsr_health_service = HSRHealthCheckService(mysql_client, pg_client) if mysql_client else None
    
    if not hsr_health_service:
        return jsonify({"error": "高铁服务未初始化"}), 500
    
    try:
        lines = hsr_health_service.get_hsr_lines()
        return jsonify({"lines": lines})
    except Exception as e:
        logging.error(f"获取高铁线路失败: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500
    
    # 获取网格小区列表和网格信息
    cells = grid_service.get_grid_cells(grid_id)
    
    from datetime import datetime, timedelta
    now = datetime.now()
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # 获取当天累计流量
    daily_traffic = grid_service.get_grid_daily_traffic(grid_id)
    
    # 获取最新1小时指标
    latest_hour_metrics = grid_service.get_grid_latest_hour_metrics(grid_id)
    
    # 时间范围显示
    time_range_display = f"当日累积（{now.strftime('%Y-%m-%d')} 00:00 至今）"
    
    # 获取仪表盘统计数据，检查该网格是否有问题标签
    grid_issues = []
    cell_stats = {
        '4g': {'no_traffic': 0, 'high_load': 0, 'total': 0},
        '5g': {'no_traffic': 0, 'high_load': 0, 'total': 0}
    }
    
    try:
        # 获取阈值参数（与网格列表页面保持一致）
        prb_threshold_4g = float(request.args.get("prb_4g", 50.0))
        prb_threshold_5g = float(request.args.get("prb_5g", 50.0))
        
        # 使用缓存获取仪表盘统计（5分钟缓存）
        cache_key = f"grid_dashboard:{prb_threshold_4g}:{prb_threshold_5g}:daily"
        dashboard_stats = cache_5m.get(
            cache_key,
            lambda: grid_service.get_dashboard_stats(
                prb_threshold_4g=prb_threshold_4g,
                prb_threshold_5g=prb_threshold_5g,
                comparison_mode='daily'
            )
        )
        
        # 检查该网格是否在流量劣化列表中
        if dashboard_stats and dashboard_stats.get('traffic_degraded_grids'):
            for grid_data in dashboard_stats['traffic_degraded_grids']:
                if grid_data['grid_id'] == grid_id:
                    grid_issues.append({
                        'type': 'traffic_degraded',
                        'label': '流量劣化',
                        'class': 'warning',
                        'detail': f"{grid_data['change_rate']:.1f}%",
                        'description': f"流量下降 {abs(grid_data['change_rate']):.1f}%"
                    })
                    break
        
        # 检查该网格是否在无流量小区增加列表中
        if dashboard_stats and dashboard_stats.get('no_traffic_increased_grids'):
            for grid_data in dashboard_stats['no_traffic_increased_grids']:
                if grid_data['grid_id'] == grid_id:
                    grid_issues.append({
                        'type': 'no_traffic_increased',
                        'label': '无流量小区增加',
                        'class': 'info',
                        'detail': f"+{grid_data['increase_rate']:.1f}%",
                        'description': f"无流量小区增加 {grid_data['increase_rate']:.1f}%"
                    })
                    break
        
        # 计算无流量小区和高负荷小区数量
        cell_stats = grid_service.get_grid_cell_stats(
            grid_id, 
            start_time=start_of_day, 
            end_time=now,
            prb_threshold_4g=prb_threshold_4g,
            prb_threshold_5g=prb_threshold_5g
        )
                    
    except Exception as e:
        logging.error(f"获取网格问题标签失败: {e}", exc_info=True)
        # 即使获取失败，也继续显示页面
    
    # 优先使用 grid_info 表的网格名称
    grid_info = cells.get('grid_info')
    grid_name = grid_info.get('grid_name') if grid_info else None
    
    return render_template(
        "grid_detail.html",
        grid_id=grid_id,
        grid_name=grid_name,
        grid_info=grid_info,
        grid_issues=grid_issues,
        cell_4g_count=len(cells.get('4g', [])),
        cell_5g_count=len(cells.get('5g', [])),
        cell_stats=cell_stats,
        daily_traffic=daily_traffic,  # 当天累计流量
        latest_hour_metrics=latest_hour_metrics,  # 最新1小时指标
        time_range_display=time_range_display,
    )


# ==================== 保障小区体检路由 ====================

@grid_bp.route("/guarantee/health_check")
@login_required
def guarantee_health_check():
    """保障小区体检页面"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 获取数据库客户端
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    
    # 初始化保障体检服务
    from services.guarantee_health_check import GuaranteeHealthCheckService
    guarantee_service = GuaranteeHealthCheckService(mysql_client, pg_client)
    
    # 获取选择的场景
    selected_scenes = request.args.getlist('scenes')
    
    try:
        # 执行体检
        health_result = guarantee_service.check_guarantee_health(selected_scenes)
        
        if 'error' in health_result:
            flash(f"体检失败: {health_result['error']}", "danger")
            return redirect(url_for("grid.grid"))
        
        # 计算健康率
        total_cells = health_result.get('total_cells', 0)
        healthy_cells = health_result.get('healthy_cells', 0)
        unhealthy_cells = health_result.get('unhealthy_cells', 0)
        healthy_rate = round(healthy_cells / total_cells * 100, 2) if total_cells > 0 else 0
        
        return render_template(
            "guarantee_health_check.html",
            check_time=health_result.get('check_time'),
            total_cells=total_cells,
            healthy_cells=healthy_cells,
            unhealthy_cells=unhealthy_cells,
            healthy_rate=healthy_rate,
            cells=health_result.get('cells', []),
            no_traffic_cells=health_result.get('no_traffic_cells', []),
            no_performance_cells=health_result.get('no_performance_cells', []),
            selected_scenes=selected_scenes
        )
        
    except Exception as e:
        logging.error(f"保障小区体检失败: {e}", exc_info=True)
        flash(f"体检失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))


@grid_bp.route("/guarantee/scenes")
@login_required
def get_guarantee_scenes():
    """获取可用的保障场景列表"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 获取数据库客户端
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    
    try:
        # 初始化保障体检服务
        from services.guarantee_health_check import GuaranteeHealthCheckService
        guarantee_service = GuaranteeHealthCheckService(mysql_client, pg_client)
        
        # 获取可用场景
        scenes = guarantee_service.get_available_scenes()
        
        return jsonify({
            'scenes': scenes
        })
        
    except Exception as e:
        logging.error(f"获取保障场景失败: {e}", exc_info=True)
        return jsonify({
            'scenes': []
        })

@grid_bp.route("/guarantee/export/health_check")
@login_required
def export_guarantee_health_check():
    """导出保障小区体检结果"""
    # 获取数据库客户端
    mysql_client = current_app.config.get('mysql_client')
    pg_client = current_app.config.get('pg_client')
    
    from services.guarantee_health_check import GuaranteeHealthCheckService
    guarantee_service = GuaranteeHealthCheckService(mysql_client, pg_client)
    
    # 获取选择的场景
    selected_scenes = request.args.getlist('scenes')
    
    try:
        # 执行体检
        health_result = guarantee_service.check_guarantee_health(selected_scenes)
        
        if 'error' in health_result:
            flash(f"体检失败: {health_result['error']}", "danger")
            return redirect(url_for("grid.grid"))
        
        # 创建Excel工作簿
        from openpyxl import Workbook
        import io
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "保障小区体检结果"
        
        # 设置表头
        headers = [
            '场景名', '小区名称', 'CGI', '制式', '健康状态', '原因',
            '是否有性能数据', '当日累计流量(GB)', '是否有告警', '告警明细',
            '前一日忙时利用率(%)', '当日忙时利用率(%)'
        ]
        ws.append(headers)
        
        # 填充数据
        for cell in health_result.get('cells', []):
            ws.append([
                cell.get('scene_name', ''),
                cell.get('celname', ''),
                cell.get('cgi', ''),
                cell.get('network_type', '').upper(),
                cell.get('status', ''),
                cell.get('reason', ''),
                cell.get('has_performance', ''),
                round(cell.get('today_traffic_gb', 0), 4),
                cell.get('has_alarm', ''),
                cell.get('alarm_details', ''),
                round(cell.get('yesterday_busy_hour_util', 0), 2),
                round(cell.get('today_busy_hour_util', 0), 2)
            ])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        from datetime import datetime
        filename = f"保障小区体检_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出保障小区体检数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))

# ==================== 网格小区体检路由 ====================

@grid_bp.route("/grid/<grid_id>/health_check")
@login_required
def grid_health_check(grid_id):
    """网格小区体检页面"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 检查服务是否可用
    if not health_check_service:
        flash("体检服务未初始化，请检查数据库配置", "warning")
        return redirect(url_for("grid.grid_detail", grid_id=grid_id))
    
    try:
        # 执行体检
        health_result = health_check_service.check_grid_health(grid_id)
        
        if 'error' in health_result:
            flash(f"体检失败: {health_result['error']}", "danger")
            return redirect(url_for("grid.grid_detail", grid_id=grid_id))
        
        # 按状态分组
        healthy_cells = [c for c in health_result['cells'] if c['status'] == 'healthy']
        unhealthy_cells = [c for c in health_result['cells'] if c['status'] == 'unhealthy']
        
        # 按原因分组不健康小区
        unhealthy_by_reason = {}
        for cell in unhealthy_cells:
            reason = cell['reason']
            if reason not in unhealthy_by_reason:
                unhealthy_by_reason[reason] = []
            unhealthy_by_reason[reason].append(cell)
        
        return render_template(
            "grid_health_check.html",
            grid_id=grid_id,
            grid_name=health_result['grid_name'],
            grid_pp=health_result.get('grid_pp', ''),
            check_time=health_result['check_time'],
            total_cells=health_result['total_cells'],
            healthy_cells_count=health_result['healthy_cells'],
            unhealthy_cells_count=health_result['unhealthy_cells'],
            healthy_rate=health_result['healthy_rate'],
            healthy_cells=healthy_cells,
            unhealthy_cells=unhealthy_cells,
            unhealthy_by_reason=unhealthy_by_reason
        )
        
    except Exception as e:
        logging.error(f"网格体检页面渲染失败: {e}", exc_info=True)
        flash(f"体检失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid_detail", grid_id=grid_id))


@grid_bp.route("/grid/health_check/all")
@admin_required
def all_grids_health_check():
    """全量网格体检页面（仅管理员）"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # 检查服务是否可用
    if not health_check_service:
        flash("体检服务未初始化，请检查数据库配置", "warning")
        return redirect(url_for("grid.grid"))
    
    try:
        # 执行全量体检（使用 30 分钟缓存，避免同一时段重复全量体检）
        logging.info("开始执行全量网格体检（支持缓存）...")
        cache_key = "grid_health_check_all"
        data = cache_30m.get(
            cache_key,
            lambda: health_check_service.check_all_grids_health_with_cells(),
        )
        health_results = data.get("results", []) if isinstance(data, dict) else (data or [])
        
        # 统计汇总
        total_grids = len(health_results)
        total_cells = sum(r['total_cells'] for r in health_results)
        total_healthy = sum(r['healthy_cells'] for r in health_results)
        total_unhealthy = sum(r['unhealthy_cells'] for r in health_results)
        overall_healthy_rate = round(total_healthy / total_cells * 100, 2) if total_cells > 0 else 0
        
        # 按健康率排序（不健康的在前）
        health_results.sort(key=lambda x: x['healthy_rate'])
        
        logging.info(f"全量网格体检完成: {total_grids}个网格, {total_cells}个小区, 健康率{overall_healthy_rate}%（结果已缓存 30 分钟）")
        
        return render_template(
            "grid_health_check_all.html",
            health_results=health_results,
            total_grids=total_grids,
            total_cells=total_cells,
            total_healthy=total_healthy,
            total_unhealthy=total_unhealthy,
            overall_healthy_rate=overall_healthy_rate,
            check_time=datetime.now()
        )
        
    except Exception as e:
        logging.error(f"全量网格体检失败: {e}", exc_info=True)
        flash(f"体检失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))


@grid_bp.route("/grid/export/health_check")
@admin_required
def export_health_check():
    """导出全量网格体检结果（包含统计和明细两个sheet）（仅管理员）"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    try:
        if not health_check_service:
            flash("体检服务未初始化", "warning")
            return redirect(url_for("grid.grid"))
        
        # 使用与全量体检相同的 30 分钟缓存，避免导出时再次全量重算
        cache_key = "grid_health_check_all"
        data = cache_30m.get(
            cache_key,
            lambda: health_check_service.check_all_grids_health_with_cells(),
        )
        if not isinstance(data, dict):
            flash("体检缓存数据不完整，请重新执行全量体检后再导出", "warning")
            return redirect(url_for("grid.all_grids_health_check"))
        
        health_results = data.get("results", [])
        cells_by_grid = data.get("cells_by_grid", {})
        
        if not health_results:
            flash("没有体检数据可导出", "warning")
            return redirect(url_for("grid.grid"))
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # ===== Sheet 1: 网格健康度统计 =====
        ws1 = wb.active
        ws1.title = "网格健康度统计"
        
        # 设置表头
        headers1 = [
            '排名', '网格ID', '网格名称', '督办标签', '网格区域', '归属单位',
            '总小区数', '健康小区数', '不健康小区数', '健康率(%)', '体检时间'
        ]
        ws1.append(headers1)
        
        # 按健康率排序（不健康的在前）
        health_results_sorted = sorted(health_results, key=lambda x: x.get('healthy_rate', 0))
        
        # 填充数据
        for i, result in enumerate(health_results_sorted, 1):
            ws1.append([
                i,
                result.get('grid_id', ''),
                result.get('grid_name', ''),
                result.get('grid_pp', ''),
                result.get('grid_area', ''),
                result.get('gird_dd', ''),
                result.get('total_cells', 0),
                result.get('healthy_cells', 0),
                result.get('unhealthy_cells', 0),
                result.get('healthy_rate', 0),
                result.get('check_time', '').strftime('%Y-%m-%d %H:%M:%S') if result.get('check_time') else ''
            ])
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 8
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 25
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 20
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 12
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 12
        ws1.column_dimensions['K'].width = 20
        
        # ===== Sheet 2: 小区明细 =====
        ws2 = wb.create_sheet(title="小区明细")
        
        # 设置表头
        headers2 = [
            '网格ID', '网格名称', '督办标签', '小区CGI', '小区名称', '制式',
            '健康状态', '原因', '有性能数据', '流量(GB)', '有告警', '告警数量', '告警明细', '最后更新时间'
        ]
        ws2.append(headers2)
        
        # 收集所有小区的详细信息（复用缓存中的 cells_by_grid，避免再次逐网格体检）
        logging.info("开始收集小区明细数据（复用缓存结果）...")
        all_cells = []
        for result in health_results_sorted:
            grid_id = result['grid_id']
            grid_cells = cells_by_grid.get(grid_id, [])
            for cell in grid_cells:
                network_type = cell.get('network_type') or ''
                # 格式化告警明细
                alarms = cell.get('alarms', [])
                alarm_details = '; '.join(
                    [f"{alarm.get('alarm_name', '')} ({alarm.get('severity', '')})" for alarm in alarms]
                ) if alarms else '无告警'
                
                all_cells.append({
                    'grid_id': grid_id,
                    'grid_name': result.get('grid_name') or '',
                    'grid_pp': result.get('grid_pp') or '',
                    'cgi': cell.get('cgi') or '',
                    'celname': cell.get('celname') or '',
                    'network_type': network_type.upper() if network_type else '',
                    'status': '正常' if cell.get('status') == 'healthy' else '不正常',
                    'reason': cell.get('reason') or '',
                    'has_performance': '是' if cell.get('has_performance') else '否',
                    'traffic_gb': round(cell.get('traffic_gb') or 0, 4),
                    'has_alarm': '是' if cell.get('has_alarm') else '否',
                    'alarm_count': cell.get('alarm_count') or 0,
                    'alarm_details': alarm_details,
                    'last_update': cell.get('last_update').strftime('%Y-%m-%d %H:%M:%S') if cell.get('last_update') else ''
                })
        
        logging.info(f"收集到 {len(all_cells)} 个小区的详细数据（来自缓存）")
        
        # 填充数据（按网格ID和小区CGI排序）
        all_cells_sorted = sorted(all_cells, key=lambda x: (x['grid_id'], x['cgi']))
        for cell in all_cells_sorted:
            ws2.append([
                cell['grid_id'],
                cell['grid_name'],
                cell['grid_pp'] or '',
                cell['cgi'],
                cell['celname'],
                cell['network_type'],
                cell['status'],
                cell['reason'],
                cell['has_performance'],
                cell['traffic_gb'],
                cell['has_alarm'],
                cell['alarm_count'],
                cell['alarm_details'],
                cell['last_update']
            ])
        
        # 设置列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 30
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 8
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 20
        ws2.column_dimensions['I'].width = 15
        ws2.column_dimensions['J'].width = 12
        ws2.column_dimensions['K'].width = 10
        ws2.column_dimensions['L'].width = 12
        ws2.column_dimensions['M'].width = 40
        ws2.column_dimensions['N'].width = 20
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"网格体检汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出体检数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))


@grid_bp.route("/grid/<grid_id>/export/health_check")
def export_grid_health_check(grid_id):
    """导出单个网格的体检详细结果（包含统计和明细两个sheet）"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    try:
        if not health_check_service:
            flash("体检服务未初始化", "warning")
            return redirect(url_for("grid.grid_detail", grid_id=grid_id))
        
        # 执行体检
        health_result = health_check_service.check_grid_health(grid_id)
        
        if 'error' in health_result:
            flash(f"体检失败: {health_result['error']}", "danger")
            return redirect(url_for("grid.grid_detail", grid_id=grid_id))
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # ===== Sheet 1: 网格健康度统计 =====
        ws1 = wb.active
        ws1.title = "网格健康度统计"
        
        # 设置表头
        headers1 = [
            '网格ID', '网格名称', '督办标签',
            '总小区数', '健康小区数', '不健康小区数', '健康率(%)', '体检时间'
        ]
        ws1.append(headers1)
        
        # 填充数据
        ws1.append([
            health_result.get('grid_id', ''),
            health_result.get('grid_name', ''),
            health_result.get('grid_pp', ''),
            health_result.get('total_cells', 0),
            health_result.get('healthy_cells', 0),
            health_result.get('unhealthy_cells', 0),
            health_result.get('healthy_rate', 0),
            health_result.get('check_time', '').strftime('%Y-%m-%d %H:%M:%S') if health_result.get('check_time') else ''
        ])
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 25
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 15
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 20
        
        # ===== Sheet 2: 小区明细 =====
        ws2 = wb.create_sheet(title="小区明细")
        
        # 设置表头
        headers2 = [
            '小区CGI', '小区名称', '制式', '健康状态', '原因',
            '有性能数据', '流量(GB)', '有告警', '告警数量', '告警明细', '最后更新时间'
        ]
        ws2.append(headers2)
        
        # 填充数据（先不健康后健康）
        cells_sorted = sorted(
            health_result['cells'], 
            key=lambda x: (0 if x.get('status') == 'unhealthy' else 1, x.get('cgi', ''))
        )
        
        for cell in cells_sorted:
            # 格式化告警明细
            alarms = cell.get('alarms', [])
            alarm_details = '; '.join([f"{alarm.get('alarm_name', '')} ({alarm.get('severity', '')})" for alarm in alarms]) if alarms else '无告警'
            
            ws2.append([
                cell.get('cgi', ''),
                cell.get('celname', ''),
                cell.get('network_type', '').upper(),
                '正常' if cell.get('status') == 'healthy' else '不正常',
                cell.get('reason', ''),
                '是' if cell.get('has_performance') else '否',
                round(cell.get('traffic_gb', 0), 4),
                '是' if cell.get('has_alarm') else '否',
                cell.get('alarm_count', 0),
                alarm_details,
                cell.get('last_update', '').strftime('%Y-%m-%d %H:%M:%S') if cell.get('last_update') else ''
            ])
        
        # 设置列宽
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 8
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 20
        ws2.column_dimensions['F'].width = 15
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 10
        ws2.column_dimensions['I'].width = 12
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 20
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"网格{grid_id}_小区体检_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出体检数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid_detail", grid_id=grid_id))


# ==================== 网格导出路由 ====================

@grid_bp.route("/grid/export/traffic_degraded")
def export_traffic_degraded():
    """导出流量劣化网格清单（包含统计和明细两个sheet）"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    try:
        if not grid_service:
            flash("网格监控服务未初始化", "warning")
            return redirect(url_for("grid.grid"))
        
        # 获取详细数据
        detailed_data = grid_service.get_traffic_degraded_details()
        
        if not detailed_data:
            flash("没有流量劣化网格数据可导出", "warning")
            return redirect(url_for("grid.grid"))
        
        # 分离统计数据和明细数据
        summary_data = [row for row in detailed_data if row.get('is_summary')]
        cell_data = [row for row in detailed_data if not row.get('is_summary')]
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # ===== Sheet 1: 网格统计 =====
        ws1 = wb.active
        ws1.title = "网格统计"
        
        # 设置表头
        headers1 = [
            '网格ID', '网格名称', '网格标签', '网格区域', '归属单位', '网格注册',
            '对比期流量', '前7天日均流量', '变化率(%)',
            '4G小区数', '5G小区数', '总小区数'
        ]
        ws1.append(headers1)
        
        # 填充数据
        for row_data in summary_data:
            ws1.append([
                row_data.get('grid_id', ''),
                row_data.get('grid_name', ''),
                row_data.get('grid_pp', ''),
                row_data.get('grid_area', ''),
                row_data.get('gird_dd', ''),
                row_data.get('grid_regration', ''),
                format_traffic(row_data.get('yesterday_traffic', 0)),
                format_traffic(row_data.get('past_7days_avg_traffic', 0)),
                round(row_data.get('change_rate', 0), 2),
                row_data.get('cell_4g_count', 0),
                row_data.get('cell_5g_count', 0),
                row_data.get('total_cells', 0)
            ])
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 25
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 20
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 18
        ws1.column_dimensions['I'].width = 12
        ws1.column_dimensions['J'].width = 12
        ws1.column_dimensions['K'].width = 12
        ws1.column_dimensions['L'].width = 12
        
        # ===== Sheet 2: 小区明细 =====
        ws2 = wb.create_sheet(title="小区明细")
        
        # 设置表头
        headers2 = [
            '网格ID', '网格名称', '网格标签', '小区CGI', '小区名称', '制式', '经度', '纬度'
        ]
        ws2.append(headers2)
        
        # 填充数据
        for row_data in cell_data:
            ws2.append([
                row_data.get('grid_id', ''),
                row_data.get('grid_name', ''),
                row_data.get('grid_pp', '') or '',  # 添加网格标签
                row_data.get('cgi', ''),
                row_data.get('celname', ''),
                row_data.get('network_type', ''),
                row_data.get('lon', ''),
                row_data.get('lat', '')
            ])
        
        # 设置列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 30
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 12
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"流量劣化网格清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出流量劣化网格数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))


@grid_bp.route("/grid/export/no_traffic_increased")
def export_no_traffic_increased():
    """导出无流量小区增加网格清单（包含统计和明细两个sheet）"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    try:
        if not grid_service:
            flash("网格监控服务未初始化", "warning")
            return redirect(url_for("grid.grid"))
        
        # 获取详细数据
        detailed_data = grid_service.get_no_traffic_increased_details()
        
        if not detailed_data:
            flash("没有无流量小区增加网格数据可导出", "warning")
            return redirect(url_for("grid.grid"))
        
        # 分离统计数据和明细数据
        summary_data = [row for row in detailed_data if row.get('is_summary')]
        cell_data = [row for row in detailed_data if not row.get('is_summary')]
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # ===== Sheet 1: 网格统计 =====
        ws1 = wb.active
        ws1.title = "网格统计"
        
        # 设置表头
        headers1 = [
            '网格ID', '网格名称', '网格标签', '网格区域', '归属单位', '网格注册',
            '昨天无流量小区数', '前7天平均无流量小区数', '增加率(%)', '总小区数'
        ]
        ws1.append(headers1)
        
        # 填充数据
        for row_data in summary_data:
            ws1.append([
                row_data.get('grid_id', ''),
                row_data.get('grid_name', ''),
                row_data.get('grid_pp', ''),
                row_data.get('grid_area', ''),
                row_data.get('gird_dd', ''),
                row_data.get('grid_regration', ''),
                row_data.get('yesterday_no_traffic_count', 0),
                round(row_data.get('past_7days_avg_no_traffic', 0), 2),
                round(row_data.get('increase_rate', 0), 2),
                row_data.get('total_cells', 0)
            ])
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 25
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 20
        ws1.column_dimensions['G'].width = 20
        ws1.column_dimensions['H'].width = 25
        ws1.column_dimensions['I'].width = 12
        ws1.column_dimensions['J'].width = 12
        
        # ===== Sheet 2: 小区明细 =====
        ws2 = wb.create_sheet(title="小区明细")
        
        # 设置表头
        headers2 = [
            '网格ID', '网格名称', '网格标签', '小区CGI', '小区名称', '制式', '流量状态', '经度', '纬度'
        ]
        ws2.append(headers2)
        
        # 填充数据
        for row_data in cell_data:
            ws2.append([
                row_data.get('grid_id', ''),
                row_data.get('grid_name', ''),
                row_data.get('grid_pp', '') or '',  # 添加网格标签
                row_data.get('cgi', ''),
                row_data.get('celname', ''),
                row_data.get('network_type', ''),
                row_data.get('has_traffic', ''),
                row_data.get('lon', ''),
                row_data.get('lat', '')
            ])
        
        # 设置列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 30
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 12
        ws2.column_dimensions['H'].width = 12
        ws2.column_dimensions['I'].width = 12
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"无流量小区增加清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出无流量小区增加数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))


@grid_bp.route("/grid/export/high_load")
def export_high_load():
    """导出高负荷小区网格清单（包含统计和明细两个sheet）"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    try:
        if not grid_service:
            flash("网格监控服务未初始化", "warning")
            return redirect(url_for("grid.grid"))
        
        # 获取仪表盘统计数据
        stats = grid_service.get_dashboard_stats()
        high_load_grids = stats.get('high_load_grids', [])
        
        if not high_load_grids:
            flash("没有高负荷小区数据可导出", "warning")
            return redirect(url_for("grid.grid"))
        
        # 获取高负荷小区明细
        high_load_cells = grid_service.get_high_load_cells_details()
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # ===== Sheet 1: 网格统计 =====
        ws1 = wb.active
        ws1.title = "网格统计"
        
        # 设置表头
        headers1 = [
            '网格ID', '网格名称', '网格标签',
            '4G高负荷小区数', '5G高负荷小区数', '总高负荷小区数', '总小区数', '高负荷占比(%)'
        ]
        ws1.append(headers1)
        
        # 填充数据
        for grid in high_load_grids:
            total_high_load = grid.get('total_high_load', 0)
            total_cells = grid.get('total_cells', 1)
            high_load_ratio = (total_high_load / total_cells * 100) if total_cells > 0 else 0
            
            ws1.append([
                grid.get('grid_id', ''),
                grid.get('grid_name', ''),
                grid.get('grid_pp', ''),
                grid.get('high_load_4g', 0),
                grid.get('high_load_5g', 0),
                total_high_load,
                total_cells,
                round(high_load_ratio, 2)
            ])
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 25
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 18
        ws1.column_dimensions['E'].width = 18
        ws1.column_dimensions['F'].width = 18
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 15
        
        # ===== Sheet 2: 小区明细 =====
        ws2 = wb.create_sheet(title="小区明细")
        
        # 设置表头
        headers2 = [
            '网格ID', '网格名称', '网格标签', '小区CGI', '小区名称', '制式', '忙时PRB利用率(%)'
        ]
        ws2.append(headers2)
        
        # 填充数据
        for cell in high_load_cells:
            ws2.append([
                cell.get('grid_id', ''),
                cell.get('grid_name', ''),
                cell.get('grid_pp', '') or '',  # 添加网格标签
                cell.get('cgi', ''),
                cell.get('celname', ''),
                cell.get('network_type', ''),
                cell.get('max_prb_util', 0)
            ])
        
        # 设置列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 30
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 20
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"高负荷小区网格清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出高负荷小区数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))


@grid_bp.route("/grid/export/fault_grids")
def export_fault_grids():
    """导出告警Top网格清单（包含统计和明细两个sheet）"""
    grid_service, health_check_service, auth_enabled = get_grid_context()
    
    # Apply login_required if auth is enabled
    if auth_enabled:
        from flask import session
        if not session.get("logged_in"):
            return redirect(url_for("login"))
    
    try:
        if not grid_service:
            flash("网格监控服务未初始化", "warning")
            return redirect(url_for("grid.grid"))
        
        # 获取仪表盘统计数据
        stats = grid_service.get_dashboard_stats()
        fault_grids_detail = stats.get('fault_grids_detail', [])
        
        if not fault_grids_detail:
            flash("没有故障网格数据可导出", "warning")
            return redirect(url_for("grid.grid"))
        
        # 获取告警明细数据
        fault_cells = []
        if grid_service.alarm_matcher:
            fault_cells = grid_service.alarm_matcher.get_fault_cells_details(performance_only=True)
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # ===== Sheet 1: 统计数据 =====
        ws1 = wb.active
        ws1.title = "网格统计"
        
        # 设置表头
        headers1 = [
            '排名', '网格ID', '网格名称', '督办标签', '故障小区数'
        ]
        ws1.append(headers1)
        
        # 填充数据
        for i, grid in enumerate(fault_grids_detail, 1):
            ws1.append([
                i,
                grid.get('grid_id', ''),
                grid.get('grid_name', ''),
                grid.get('grid_pp', ''),
                grid.get('fault_count', 0)
            ])
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 8
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 25
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['E'].width = 15
        
        # ===== Sheet 2: 小区明细 =====
        ws2 = wb.create_sheet(title="小区明细")
        
        # 设置表头
        headers2 = [
            '网格ID', '网格名称', '网格标签', '小区CGI', '小区名称', '告警名称', '告警数量'
        ]
        ws2.append(headers2)
        
        # 填充数据（按网格ID排序）
        fault_cells_sorted = sorted(fault_cells, key=lambda x: x.get('grid_id', ''))
        for cell in fault_cells_sorted:
            ws2.append([
                cell.get('grid_id', ''),
                cell.get('grid_name', ''),
                cell.get('grid_pp', '') or '',  # 确保None转换为空字符串
                cell.get('cgi', ''),
                cell.get('celname', ''),
                cell.get('alarm_names', ''),
                cell.get('alarm_count', 0)
            ])
        
        # 设置列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 30
        ws2.column_dimensions['E'].width = 25
        ws2.column_dimensions['F'].width = 50
        ws2.column_dimensions['G'].width = 12
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"告警Top网格清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"导出告警Top网格数据失败: {e}", exc_info=True)
        flash(f"导出失败: {str(e)}", "danger")
        return redirect(url_for("grid.grid"))
